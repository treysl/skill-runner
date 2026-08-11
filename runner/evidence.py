from __future__ import annotations

import hashlib
import json
import subprocess
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from runner.config import OUTPUT_DIR, SKILL_PACKAGE

MANIFEST_DIR = OUTPUT_DIR / "run-manifests"


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest().upper()


def _git_state() -> dict[str, Any]:
    try:
        commit = subprocess.run(
            ["git", "rev-parse", "HEAD"],
            capture_output=True,
            text=True,
            check=True,
        ).stdout.strip()
        dirty = bool(
            subprocess.run(
                ["git", "status", "--porcelain"],
                capture_output=True,
                text=True,
                check=True,
            ).stdout.strip()
        )
        return {"commit": commit, "dirty": dirty}
    except (OSError, subprocess.SubprocessError):
        return {"commit": None, "dirty": None}


def _workbook_summary(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        dimensions = {
            name: {
                "rows": workbook[name].max_row,
                "columns": workbook[name].max_column,
            }
            for name in workbook.sheetnames
        }
        return {
            "worksheet_count": len(workbook.sheetnames),
            "worksheets": workbook.sheetnames,
            "worksheet_dimensions": dimensions,
        }
    finally:
        workbook.close()


def _artifact(path: Path, *, workbook: bool = False) -> dict[str, Any]:
    details: dict[str, Any] = {
        "filename": path.name,
        "size_bytes": path.stat().st_size,
        "sha256": sha256_file(path),
    }
    if workbook:
        details.update(_workbook_summary(path))
    return details


def _write_manifest(payload: dict[str, Any], path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(json.dumps(payload, indent=2, sort_keys=True), encoding="utf-8")
    temporary.replace(path)
    return path


def write_success_manifest(
    *,
    input_file: Path,
    output_file: Path,
    inspection: dict[str, Any],
    config: dict[str, Any],
    started_at: datetime,
    elapsed_seconds: float,
    manifest_dir: Path | None = None,
) -> Path:
    finished_at = datetime.now(timezone.utc)
    payload = {
        "schema_version": 1,
        "status": "success",
        "started_at_utc": started_at.astimezone(timezone.utc).isoformat(),
        "finished_at_utc": finished_at.isoformat(),
        "elapsed_seconds": round(elapsed_seconds, 3),
        "input": _artifact(input_file),
        "inspection": {
            "row_count": inspection.get("row_count"),
            "column_count": len(inspection.get("columns") or []),
            "missing_required_columns": inspection.get("missing_required_columns") or [],
            "branches": inspection.get("branches") or [],
            "divisions": inspection.get("divisions") or [],
        },
        "configuration": config,
        "output": _artifact(output_file, workbook=True),
        "skill_package": _artifact(SKILL_PACKAGE),
        "code": _git_state(),
        "warnings": [],
    }
    destination = (manifest_dir or MANIFEST_DIR) / f"{output_file.stem}.manifest.json"
    return _write_manifest(payload, destination)


def write_failure_manifest(
    *,
    started_at: datetime,
    elapsed_seconds: float,
    error: Exception,
    input_file: Path | None = None,
    inspection: dict[str, Any] | None = None,
    config: dict[str, Any] | None = None,
    manifest_dir: Path | None = None,
) -> Path:
    finished_at = datetime.now(timezone.utc)
    payload: dict[str, Any] = {
        "schema_version": 1,
        "status": "failure",
        "started_at_utc": started_at.astimezone(timezone.utc).isoformat(),
        "finished_at_utc": finished_at.isoformat(),
        "elapsed_seconds": round(elapsed_seconds, 3),
        "input": _artifact(input_file) if input_file and input_file.is_file() else None,
        "inspection": inspection,
        "configuration": config,
        "output": None,
        "skill_package": _artifact(SKILL_PACKAGE),
        "code": _git_state(),
        "warnings": [],
        "error": {"type": type(error).__name__, "message": str(error)},
    }
    stamp = finished_at.strftime("%Y%m%d_%H%M%S_%f")
    destination = (manifest_dir or MANIFEST_DIR) / f"failed_run_{stamp}.manifest.json"
    return _write_manifest(payload, destination)
