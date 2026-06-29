#!/usr/bin/env python3
"""
CIP Report Builder — The Herring Group
Canonical build script for the Construction In Process report.

Usage:
    python3 build_cip_report.py <input_file> <output_file> [options]

Options:
    --client-name NAME          Client display name (required)
    --branch BRANCH             Branch filter (repeatable; omit for all)
    --division DIV              Division filter (repeatable; required)
    --opp-status STATUS         (REMOVED — now auto-filters to statuses containing "Won")
    --original-opp-cutoff DATE  Won Date cutoff for Original Opp flag (YYYY-MM-DD, required)
    --completed-months N        Lookback window for completed jobs (default: 12)
    --user NAME                 User name for Version History (default: Trey)
    --change-note NOTE          Change description for Version History
    --logo PATH                 Path to THG logo PNG
    --no-logo                   Skip logo insertion
    --min-est-revenue AMOUNT    Minimum Revenue Estimated threshold (exclude opps below; default: 0)

All values are computed in Python and written as static numbers (no in-cell formulas).
Formula recalculation is not required.
"""

import argparse
import sys
import os

# ── Force UTF-8 on this script's own output streams ──────────────────────────
# Status lines below emit non-ASCII (✅, em dashes, → arrows, ═ rules, and
# interpolated client/column names). A Windows console defaulting to a legacy
# code page (e.g. cp1252) would raise UnicodeEncodeError and abort the build.
# reconfigure() exists on TextIOWrapper in Python 3.7+; errors="replace" keeps
# the run alive even if an exotic glyph slips through. Output files themselves
# (.xlsx via openpyxl, .xlsx input via pandas) are already UTF-8 internally and
# unaffected by OS locale — this only hardens stdout/stderr.
for _stream in (sys.stdout, sys.stderr):
    try:
        _stream.reconfigure(encoding="utf-8", errors="replace")
    except (AttributeError, ValueError):
        pass  # stream isn't a reconfigurable TextIOWrapper (e.g. redirected) — skip

import numpy as np
import pandas as pd
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AbsoluteAnchor
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

try:
    import pytz
except ImportError:
    print("ERROR: pytz required. Install with: pip install pytz --break-system-packages", file=sys.stderr)
    sys.exit(1)

try:
    from PIL import Image as PILImage
except ImportError:
    print("ERROR: Pillow required. Install with: pip install Pillow --break-system-packages", file=sys.stderr)
    sys.exit(1)


# ═══════════════════════════════════════════════════════════════════════════════
# CONSTANTS — THG Brand (from thg-brand-guidelines + thg-report-standards Std 5)
# ═══════════════════════════════════════════════════════════════════════════════

NAVY = "082544"
OLIVE = "807E34"
MANGO = "F07F42"
HERON = "434E4F"
WHITE_HEX = "FFFFFF"
ZEBRA_HEX = "F5F5F5"

# Conditional formatting colors (CIP-specific)
CF_GREEN = "C6EFCE"
CF_RED = "F4CCCC"
CF_ORANGE = "F4B084"
CF_YELLOW = "FFF2CC"

# Fills
FILL_NAVY = PatternFill("solid", fgColor=NAVY)
FILL_OLIVE = PatternFill("solid", fgColor=OLIVE)
FILL_MANGO = PatternFill("solid", fgColor=MANGO)
FILL_HERON = PatternFill("solid", fgColor=HERON)
FILL_WHITE = PatternFill("solid", fgColor=WHITE_HEX)
FILL_ZEBRA = PatternFill("solid", fgColor=ZEBRA_HEX)
FILL_HEADER = PatternFill("solid", fgColor=NAVY)
FILL_SUBTOTAL = PatternFill("solid", fgColor="D9E1F2")
FILL_GREEN = PatternFill("solid", fgColor=CF_GREEN)
FILL_RED = PatternFill("solid", fgColor=CF_RED)
FILL_ORANGE = PatternFill("solid", fgColor=CF_ORANGE)
FILL_YELLOW = PatternFill("solid", fgColor=CF_YELLOW)
FILL_NOT_STARTED = PatternFill("solid", fgColor="BDD7EE")  # Light blue highlight for projects not yet started

# Fonts (thg-report-standards Std 5: Merriweather headings, Roboto body)
def font_heading(size=12, bold=True, color=NAVY):
    return Font(name="Merriweather", size=size, bold=bold, color=color)

def font_body(size=10, bold=False, italic=False, color="000000"):
    return Font(name="Roboto", size=size, bold=bold, italic=italic, color=color)

FONT_KPI_VAL = Font(name="Roboto", size=16, bold=True, color=NAVY)
FONT_KPI_LABEL = Font(name="Roboto", size=8, color=HERON)
FONT_CF_NOTE = Font(name="Roboto", size=8, italic=True, color=HERON)

# Alignments
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_LEFT_NOWRAP = Alignment(horizontal="left", vertical="center", wrap_text=False)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
ALIGN_PCT = Alignment(horizontal="center", vertical="center")
ALIGN_VCENTER = Alignment(vertical="center")
ALIGN_VCENTER_NOWRAP = Alignment(vertical="center", wrap_text=False)

# Number formats (thg-report-standards Std 5)
FMT_DOLLAR = '_($* #,##0_)'
FMT_PCT = '0.0%'
FMT_HOURS = '#,##0.0'
FMT_COUNT = '#,##0'
FMT_DATE = 'MM/DD/YYYY'

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

# Default buckets for Complete Overview
DEFAULT_GM_BUCKETS = [
    ("<30%", float("-inf"), 0.30), ("30%-35%", 0.30, 0.35), ("35%-40%", 0.35, 0.40),
    ("40%-45%", 0.40, 0.45), ("45%-50%", 0.45, 0.50), ("50%-60%", 0.50, 0.60),
    ("60%-70%", 0.60, 0.70), (">70%", 0.70, float("inf")),
]
DEFAULT_REV_BUCKETS = [
    ("$0-$5K", 0, 5000), ("$5K-$10K", 5000, 10000), ("$10K-$20K", 10000, 20000),
    ("$20K-$50K", 20000, 50000), ("$50K-$100K", 50000, 100000),
    ("$100K-$200K", 100000, 200000), ("$200K-$300K", 200000, 300000),
    (">$300K", 300000, float("inf")),
]

# Tab colors
TAB_COLORS = {
    "Proprietary Disclosure": NAVY,
    "Dashboard": MANGO,
    "In Process": OLIVE,
    "In Process By Property": OLIVE,
    "Completed": NAVY,
    "Completed Dashboard View": MANGO,
    "Complete Overview": MANGO,
    "Version History": HERON,
}

# ═══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def safe_div(num, den):
    """Return 0 when denominator is zero, null, or NaN. Works on scalars and Series."""
    if isinstance(num, (pd.Series, np.ndarray)):
        return np.where((den == 0) | pd.isna(den) | pd.isna(num), 0, num / den)
    if den is None or den == 0 or (isinstance(den, float) and np.isnan(den)):
        return 0
    if num is None or (isinstance(num, float) and np.isnan(num)):
        return 0
    return num / den


def _fmt_dt(dt, with_time=False):
    """Cross-platform M/D/YYYY (optionally with h:MM AM/PM) — avoids Linux-only %-m/%-d/%-I."""
    s = f"{dt.month}/{dt.day}/{dt.year}"
    if with_time:
        hour12 = dt.hour % 12 or 12
        ampm = "AM" if dt.hour < 12 else "PM"
        s += f" {hour12}:{dt.minute:02d} {ampm}"
    return s


def run_date_str():
    tz = pytz.timezone("America/Chicago")
    now = datetime.now(tz)
    suffix = "CDT" if now.dst() else "CST"
    return f"{_fmt_dt(now, with_time=True)} US {suffix}"


def add_logo(ws, logo_path):
    if not logo_path or not os.path.exists(logo_path):
        return
    with PILImage.open(logo_path) as img:
        nw, nh = img.size
    target_w = 100
    logo_h = round(nh * target_w / nw)
    logo = XLImage(logo_path)
    logo.width = target_w
    logo.height = logo_h
    anchor = AbsoluteAnchor(
        pos=XDRPoint2D(pixels_to_EMU(4), pixels_to_EMU(3)),
        ext=XDRPositiveSize2D(pixels_to_EMU(target_w), pixels_to_EMU(logo_h)),
    )
    logo.anchor = anchor
    ws.add_image(logo)
    ws.row_dimensions[1].height = round(logo_h * 0.75) + 4


def apply_cell(ws, row, col, value, font=None, fill=None, alignment=None,
               number_format=None, border=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font: cell.font = font
    if fill: cell.fill = fill
    if alignment: cell.alignment = alignment
    if number_format: cell.number_format = number_format
    if border: cell.border = border
    return cell


# ═══════════════════════════════════════════════════════════════════════════════
# DATA PIPELINE
# ═══════════════════════════════════════════════════════════════════════════════

EST_COST_COLS = [
    'Labor Cost Estimated', 'Material Cost Estimated', 'Sub Cost Estimated',
    'Equipment Cost Estimated', 'Other Cost Estimated',
]

SUM_ACTUAL_COLS = [
    'Earned Revenue', 'Revenue Estimated', 'Invoiced Revenue',
    'Labor Hours Actual', 'Labor Hours Estimated', 'Future Scheduled Hours',
    'Labor Cost Actual', 'Material Cost Actual', 'Sub Cost Actual',
    'Equipment Cost Actual', 'Other Cost Actual',
]

SUM_EST_COLS = [c + '.1' for c in EST_COST_COLS]


def build_opp_master(raw):
    om = raw.copy()
    om = om[om['Opportunity Type'] == 'Work Order']
    om = om[om['Job Status'] != 'Canceled']
    om['_rev_str'] = om['Revision #'].astype(str)
    om = om[om['_rev_str'].notna() & (om['_rev_str'] != '') & (om['_rev_str'] != 'nan')]
    om = om[~om['_rev_str'].str.contains(' ', na=False)]
    om = om[~om['Invoice Type'].str.contains('T&M', na=False)]
    om['Master Opportunity Name'] = om['Master Opportunity Name'].fillna(om['Opportunity Name'])
    mask = om['Master Opportunity Name'] == ''
    om.loc[mask, 'Master Opportunity Name'] = om.loc[mask, 'Opportunity Name']

    opp_master = om.groupby('Opportunity #').agg({
        'Master Opportunity Name': 'first',
        'Company Name': 'first',
        'Sales Rep': 'first',
        'Operations Mgr Name': 'first',
        'Start Date': 'min',
    }).reset_index()
    opp_master = opp_master[opp_master['Opportunity #'].notna()]
    opp_master.columns = ['Opportunity #', 'Master Opp Name', 'Company Name',
                          'Sales Rep', 'Ops Mgr', 'Start Date OM']
    return opp_master


def apply_original_opp_flag(dq, cutoff):
    dq['_rev_str'] = dq['Revision #'].astype(str)
    def _flag(row):
        won = row['Won Date']
        rev = row['_rev_str']
        if pd.isna(won):
            return 1
        if won > cutoff:
            return 1
        if '.' not in rev or rev.endswith('.0'):
            return 1
        return 0
    dq['_orig_opp'] = dq.apply(_flag, axis=1)
    for c in EST_COST_COLS:
        dq[c + '.1'] = dq[c].fillna(0) * dq['_orig_opp']
    return dq


def resolve_completed_window(kind, today, completed_start=None, completed_end=None):
    """Completed-tab window. Ends on the CURRENT DAY (inclusive), not month-end."""
    if completed_start:
        start = pd.Timestamp(completed_start)
        end = (pd.Timestamp(completed_end) + pd.Timedelta(days=1)) if completed_end \
            else (today.normalize() + pd.Timedelta(days=1))
        return start, end, "custom"
    end = today.normalize() + pd.Timedelta(days=1)  # exclusive bound → includes today
    if kind == "this_month":
        start = today.normalize().replace(day=1)
    elif kind == "last_30_days":
        start = today.normalize() - pd.Timedelta(days=30)
    elif kind == "ytd":
        start = today.normalize().replace(month=1, day=1)
    else:
        start = today.normalize().replace(day=1)
    return start, end, kind


def resolve_overview_window(kind, today):
    """Complete Overview window — independent of the Completed-tab window."""
    if kind == "last_year":
        y = today.year - 1
        return pd.Timestamp(y, 1, 1), pd.Timestamp(y + 1, 1, 1), kind
    # last_12_complete_months: 12 full calendar months ending the month before current
    first_of_cur = today.normalize().replace(day=1)
    start = first_of_cur - pd.DateOffset(months=12)
    return start, first_of_cur, kind


def build_data_queries(raw, opp_master, divisions, branches, cutoff,
                       completed_range="this_month", overview_range="last_12_complete_months",
                       min_est_revenue=0, completed_start=None, completed_end=None):
    dq = raw.copy()
    dq = dq[dq['Opportunity Type'] == 'Work Order']
    dq = dq[~dq['Invoice Type'].str.contains('T&M', na=False)]
    dq = apply_original_opp_flag(dq, cutoff)
    dq = dq[dq['Opportunity Status Name'].str.contains('won', case=False, na=False)]
    matched_statuses = dq['Opportunity Status Name'].unique().tolist()
    print(f"  Opp Status filter (contains 'won'): matched {matched_statuses}")
    dq = dq[dq['Division'].isin(divisions)]
    if branches:
        dq = dq[dq['Branch'].isin(branches)]

    # Join Opp Master
    dq = dq.merge(opp_master, on='Opportunity #', how='left')
    dq['Opp Name'] = dq['Master Opp Name'].fillna(dq['Opportunity Name'])
    dq['Company Name'] = dq['Company Name_y'].fillna(dq['Company Name_x'])
    dq['Sales Rep'] = dq['Sales Rep_y'].fillna(dq['Sales Rep_x'])
    dq['Ops Mgr'] = dq['Ops Mgr'].fillna(dq['Operations Mgr Name'])
    dq['Start Date Final'] = dq['Start Date OM'].fillna(dq['Start Date'])

    agg_sum = {c: 'sum' for c in SUM_ACTUAL_COLS + SUM_EST_COLS}

    # In Process — group by Opp #
    ip_raw = dq[dq['Job Status'].str.contains('in p', case=False, na=False)]
    ip_statuses = ip_raw['Job Status'].unique().tolist()
    print(f"  Job Status filter (contains 'in p'): matched {ip_statuses}")
    if len(ip_raw) > 0:
        ip_grouped = ip_raw.groupby('Opportunity #').agg({
            'Property Name': 'first', 'Opp Name': 'first', 'Company Name': 'first',
            'Sales Rep': 'first', 'Ops Mgr': 'first', 'Start Date Final': 'min',
            'Job Status': 'first', 'Branch': 'first', 'Division': 'first',
            'Invoice Type': 'first', 'Opportunity Status Name': 'first',
            **agg_sum,
        }).reset_index()
    else:
        ip_grouped = pd.DataFrame()

    # In Process By Property
    bp_keys = ['Property Name', 'Company Name', 'Job Status', 'Division',
               'Opportunity Status Name', 'Branch']
    if len(ip_raw) > 0:
        ip_byprop = ip_raw.groupby(bp_keys, dropna=False).agg({'Start Date Final': 'min', **agg_sum}).reset_index()
    else:
        ip_byprop = pd.DataFrame()

    # Completed candidate rows (status only) — windowed twice below
    comp_all = dq[dq['Job Status'].str.contains('complete', case=False, na=False)]
    comp_statuses = comp_all['Job Status'].unique().tolist()
    print(f"  Job Status filter (contains 'complete'): matched {comp_statuses}")

    comp_agg = {
        'Property Name': 'first', 'Opp Name': 'first', 'Company Name': 'first',
        'Sales Rep': 'first', 'Ops Mgr': 'first', 'Start Date Final': 'min',
        'Job Status': 'first', 'Branch': 'first', 'Division': 'first',
        'Invoice Type': 'first', 'Opportunity Status Name': 'first',
        'Oppty Complete Date': 'first', **agg_sum,
    }

    def _window_completed(start, end):
        sub = comp_all[(comp_all['Oppty Complete Date'] >= start) &
                       (comp_all['Oppty Complete Date'] < end)]
        if len(sub) == 0:
            return pd.DataFrame()
        return sub.groupby('Opportunity #').agg(comp_agg).reset_index()

    today = pd.Timestamp.now()

    # Completed tab window (to current day)
    c_start, c_end, c_kind = resolve_completed_window(
        completed_range, today, completed_start, completed_end)
    comp_grouped = _window_completed(c_start, c_end)
    print(f"  Completed-tab window [{c_kind}]: "
          f"{c_start.strftime('%m/%d/%Y')} to {(c_end - pd.Timedelta(days=1)).strftime('%m/%d/%Y')}")

    # Complete Overview window (separate)
    o_start, o_end, o_kind = resolve_overview_window(overview_range, today)
    overview_grouped = _window_completed(o_start, o_end)
    print(f"  Complete Overview window [{o_kind}]: "
          f"{o_start.strftime('%m/%d/%Y')} to {(o_end - pd.Timedelta(days=1)).strftime('%m/%d/%Y')}")

    # Pre-threshold IP copy for Dashboard backlog-by-division (includes ALL opps)
    ip_grouped_all = ip_grouped.copy() if len(ip_grouped) > 0 else pd.DataFrame()

    # Apply minimum estimated revenue filter (after grouping so it's on Opp-level totals)
    if min_est_revenue > 0:
        before_ip = len(ip_grouped)
        before_comp = len(comp_grouped)
        if len(ip_grouped) > 0:
            ip_grouped = ip_grouped[ip_grouped['Revenue Estimated'] > min_est_revenue].reset_index(drop=True)
        if len(ip_byprop) > 0:
            ip_byprop = ip_byprop[ip_byprop['Revenue Estimated'] > min_est_revenue].reset_index(drop=True)
        if len(comp_grouped) > 0:
            comp_grouped = comp_grouped[comp_grouped['Revenue Estimated'] > min_est_revenue].reset_index(drop=True)
        if len(overview_grouped) > 0:
            overview_grouped = overview_grouped[overview_grouped['Revenue Estimated'] > min_est_revenue].reset_index(drop=True)
        print(f"  Min Est Revenue filter (>${min_est_revenue:,.0f}): "
              f"IP {before_ip}→{len(ip_grouped)}, Comp {before_comp}→{len(comp_grouped)} "
              f"(backlog-by-division uses pre-threshold IP set of {len(ip_grouped_all)})")

    return ip_grouped, ip_byprop, comp_grouped, overview_grouped, ip_grouped_all, dq, c_start


def calc_columns(df):
    if len(df) == 0:
        return df
    df = df.copy()
    df['Act Cost $'] = (df['Labor Cost Actual'].fillna(0) + df['Material Cost Actual'].fillna(0) +
                        df['Sub Cost Actual'].fillna(0) + df['Equipment Cost Actual'].fillna(0) +
                        df['Other Cost Actual'].fillna(0))
    df['Estimated Cost $'] = (df['Labor Cost Estimated.1'].fillna(0) + df['Material Cost Estimated.1'].fillna(0) +
                              df['Sub Cost Estimated.1'].fillna(0) + df['Equipment Cost Estimated.1'].fillna(0) +
                              df['Other Cost Estimated.1'].fillna(0))
    er = df['Earned Revenue'].fillna(0)
    re = df['Revenue Estimated'].fillna(0)
    df['Actual GM%'] = safe_div(er - df['Act Cost $'], er)
    df['Estimated GM%'] = safe_div(re - df['Estimated Cost $'], re)
    df['Rev % Completed'] = safe_div(er, re)
    df['Invoice %'] = safe_div(df['Invoiced Revenue'].fillna(0), re)
    df['Total Act/Est Cost'] = safe_div(df['Act Cost $'], df['Estimated Cost $'])
    df['Labor Hrs Act/Est'] = safe_div(df['Labor Hours Actual'].fillna(0), df['Labor Hours Estimated'].fillna(0))
    df['Labor $ Act/Est'] = safe_div(df['Labor Cost Actual'].fillna(0), df['Labor Cost Estimated.1'].fillna(0))
    df['Materials Act/Est'] = safe_div(df['Material Cost Actual'].fillna(0), df['Material Cost Estimated.1'].fillna(0))
    df['Sub Act/Est'] = safe_div(df['Sub Cost Actual'].fillna(0), df['Sub Cost Estimated.1'].fillna(0))
    df['Equip Act/Est'] = safe_div(df['Equipment Cost Actual'].fillna(0), df['Equipment Cost Estimated.1'].fillna(0))
    df['Other Act/Est'] = safe_div(df['Other Cost Actual'].fillna(0), df['Other Cost Estimated.1'].fillna(0))
    df['Actual+Sched Hrs'] = df['Labor Hours Actual'].fillna(0) + df['Future Scheduled Hours'].fillna(0)
    _overage = df['Actual+Sched Hrs'] - df['Labor Hours Estimated'].fillna(0)
    # Only surface positive overages — unscheduled hours don't imply the job will come under.
    df['Potential Hr Overage'] = np.where(_overage > 0, _overage, np.nan)
    lha = df['Labor Hours Actual'].fillna(0)
    lhe = df['Labor Hours Estimated'].fillna(0)
    df['Hours Remain'] = np.where(lha < lhe, lhe - lha, 0)
    sca = df['Sub Cost Actual'].fillna(0)
    sce = df['Sub Cost Estimated.1'].fillna(0)
    df['Subs Remain'] = np.where(sca < sce, sce - sca, 0)
    mca = df['Material Cost Actual'].fillna(0)
    mce = df['Material Cost Estimated.1'].fillna(0)
    df['Mat Remain'] = np.where(mca < mce, mce - mca, 0)
    df['Backlog'] = re - er
    return df


def _finalize_completed(df):
    if len(df) == 0:
        return df
    df['Actual Gross Profit'] = df['Earned Revenue'] - df['Act Cost $']
    df['Est Gross Profit'] = df['Revenue Estimated'] - df['Estimated Cost $']
    df['Year-Month'] = df['Oppty Complete Date'].apply(
        lambda d: d.strftime('%Y-%m') if pd.notna(d) else '')
    df['Complete Year'] = df['Oppty Complete Date'].apply(
        lambda d: d.year if pd.notna(d) else '')
    df['Complete Quarter'] = df['Oppty Complete Date'].apply(
        lambda d: f"Q{(d.month-1)//3+1}" if pd.notna(d) else '')
    return df.sort_values('Earned Revenue', ascending=False).reset_index(drop=True)


def finalize_datasets(ip_grouped, ip_byprop, comp_grouped, overview_grouped=None,
                      ip_grouped_all=None):
    ip_grouped = calc_columns(ip_grouped)
    ip_byprop = calc_columns(ip_byprop)
    comp_grouped = calc_columns(comp_grouped)
    if overview_grouped is not None:
        overview_grouped = calc_columns(overview_grouped)
    if ip_grouped_all is not None:
        ip_grouped_all = calc_columns(ip_grouped_all)

    if len(ip_grouped) > 0:
        ip_grouped['Year-Month'] = ip_grouped['Start Date Final'].apply(
            lambda d: d.strftime('%Y-%m') if pd.notna(d) else '')
        ip_grouped = ip_grouped.sort_values('Revenue Estimated', ascending=False).reset_index(drop=True)

    if len(ip_byprop) > 0:
        ip_byprop['Year-Month'] = ip_byprop['Start Date Final'].apply(
            lambda d: d.strftime('%Y-%m') if pd.notna(d) else '')
        ip_byprop = ip_byprop.sort_values('Revenue Estimated', ascending=False).reset_index(drop=True)

    comp_grouped = _finalize_completed(comp_grouped)
    if overview_grouped is not None:
        overview_grouped = _finalize_completed(overview_grouped)

    return ip_grouped, ip_byprop, comp_grouped, overview_grouped, ip_grouped_all


# ═══════════════════════════════════════════════════════════════════════════════
# WORKBOOK BUILDERS
# ═══════════════════════════════════════════════════════════════════════════════

def build_proprietary_disclosure(wb, client_name, logo_path, run_date, branches, divisions,
                                  completed_label, overview_label, sub_margin=0.281):
    ws = wb.active
    ws.title = "Proprietary Disclosure"
    ws.sheet_properties.tabColor = TAB_COLORS["Proprietary Disclosure"]
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 120

    add_logo(ws, logo_path)
    apply_cell(ws, 1, 2, f"{client_name} — Construction In Process Report",
               font=font_heading(14, True, NAVY), alignment=ALIGN_VCENTER)

    ws.merge_cells('A2:B2')
    apply_cell(ws, 2, 1,
        "CONFIDENTIAL — This report is prepared exclusively for the use of the client "
        "and The Herring Group. Do not distribute without written consent.",
        font=font_body(10, italic=True, color=HERON), alignment=ALIGN_LEFT)

    ws.merge_cells('A3:B3')
    apply_cell(ws, 3, 1, f"Run Date: {run_date}", font=font_body(10, color=HERON), alignment=ALIGN_LEFT)

    ws.merge_cells('A4:B4')
    apply_cell(ws, 4, 1, "Report Filters Applied", font=font_heading(11, True, NAVY))

    branch_str = "All" if not branches else ", ".join(branches)
    div_str = ", ".join(divisions)
    ws.merge_cells('A5:B5')
    apply_cell(ws, 5, 1,
        f"Opportunity Type = Work Order | Invoice Type excludes T&M | "
        f"Revision # valid | Branches: {branch_str} | Divisions: {div_str} | "
        f"Opp Status: contains 'Won' | Job Status: contains 'In P…' / 'Complete' | "
        f"Completed tab window: {completed_label} | Complete Overview window: {overview_label} | "
        f"Complete Overview sub margin assumption: {sub_margin:.1%}",
        font=font_body(9, color=HERON), alignment=ALIGN_LEFT)
    ws.row_dimensions[5].height = 34

    ws.merge_cells('A7:B7')
    apply_cell(ws, 7, 1, "Purpose", font=font_heading(11, True, NAVY))

    ws.merge_cells('A8:B8')
    apply_cell(ws, 8, 1,
        "This report provides a comprehensive view of active and recently completed "
        "construction/enhancement projects. It surfaces budget overruns, invoicing gaps, "
        "margin erosion, labor concerns, and backlog visibility to support proactive "
        "project management and financial oversight.",
        font=font_body(10), alignment=ALIGN_LEFT)
    ws.row_dimensions[8].height = 34

    ws.merge_cells('A10:B10')
    apply_cell(ws, 10, 1, "Helpful Information", font=font_heading(11, True, NAVY))

    # In Process
    ws.merge_cells('A11:B11')
    apply_cell(ws, 11, 1, "In Process:", font=font_body(10, bold=True))

    ws.merge_cells('A12:B12')
    apply_cell(ws, 12, 1, "The In Process tab answers questions like:", font=font_body(10))

    ip_questions = [
        "• Is a project on budget?",
        "• Have sufficient dollars been invoiced?",
        "• Are the tickets missing any materials or subs costs?",
    ]
    for i, q in enumerate(ip_questions, start=13):
        ws.merge_cells(f'A{i}:B{i}')
        apply_cell(ws, i, 1, q, font=font_body(10), alignment=ALIGN_LEFT)

    # Complete
    r = 13 + len(ip_questions) + 1  # blank row then next section
    ws.merge_cells(f'A{r}:B{r}')
    apply_cell(ws, r, 1, "Complete:", font=font_body(10, bold=True))
    r += 1
    ws.merge_cells(f'A{r}:B{r}')
    apply_cell(ws, r, 1,
        "This tab can be used to evaluate which projects were winners and which projects "
        "were not. That type of review is often helpful to improve performance moving forward.",
        font=font_body(10), alignment=ALIGN_LEFT)

    # Overview
    r += 2
    ws.merge_cells(f'A{r}:B{r}')
    apply_cell(ws, r, 1, "Overview:", font=font_body(10, bold=True))
    r += 1
    ws.merge_cells(f'A{r}:B{r}')
    apply_cell(ws, r, 1,
        "This tab breaks down completed jobs based on gross margin brackets and revenue "
        "brackets to help identify high levels of concentration.",
        font=font_body(10), alignment=ALIGN_LEFT)

    # Backlog
    r += 2
    ws.merge_cells(f'A{r}:B{r}')
    apply_cell(ws, r, 1, "Backlog:", font=font_body(10, bold=True))
    r += 1
    ws.merge_cells(f'A{r}:B{r}')
    apply_cell(ws, r, 1,
        "Backlog is the amount of work that has been sold but not performed yet. It's "
        "recommended to look at the backlog for work orders on a monthly basis to see how "
        "much revenue and hours are scheduled for the future.",
        font=font_body(10), alignment=ALIGN_LEFT)


def write_kpi(ws, row, col, label, value, fmt_type='dollar'):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = FONT_KPI_VAL
    cell.alignment = ALIGN_CENTER
    if fmt_type == 'pct':
        cell.number_format = FMT_PCT
    elif fmt_type == 'count':
        cell.number_format = FMT_COUNT
    else:
        cell.number_format = FMT_DOLLAR
    ws.cell(row=row+1, column=col, value=label).font = FONT_KPI_LABEL
    ws.cell(row=row+1, column=col).alignment = ALIGN_CENTER


def write_dash_table(ws, start_row, title, headers, data, col_formats):
    ncols = len(headers)
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=ncols)
    apply_cell(ws, start_row, 1, title, font=font_heading(10, True, WHITE_HEX),
               fill=FILL_NAVY, alignment=ALIGN_CENTER)

    hr = start_row + 1
    for ci, h in enumerate(headers, 1):
        apply_cell(ws, hr, ci, h, font=font_body(9, bold=True, color=WHITE_HEX),
                   fill=FILL_HERON, alignment=ALIGN_CENTER, border=THIN_BORDER)

    for ri, row_data in enumerate(data):
        r = hr + 1 + ri
        fill = FILL_ZEBRA if ri % 2 == 0 else FILL_WHITE
        for ci, val in enumerate(row_data):
            cf = col_formats[ci] if ci < len(col_formats) else 'text'
            if cf == 'dollar':
                nf, al = FMT_DOLLAR, ALIGN_RIGHT
            elif cf in ('pct', 'ratio'):
                nf, al = FMT_PCT, ALIGN_PCT
            elif cf == 'hours':
                nf, al = FMT_HOURS, ALIGN_LEFT_NOWRAP
            else:
                nf, al = None, ALIGN_LEFT
            apply_cell(ws, r, ci+1, val, font=font_body(9), fill=fill,
                       alignment=al, number_format=nf, border=THIN_BORDER)

    return hr + 1 + len(data) + 1


def build_dashboard(wb, client_name, logo_path, run_date, ip_grouped, comp_grouped,
                    divisions, completed_label, ip_grouped_all=None):
    ws = wb.create_sheet("Dashboard")
    ws.sheet_properties.tabColor = TAB_COLORS["Dashboard"]
    ws.sheet_view.showGridLines = False
    for c in range(1, 5):
        ws.column_dimensions[get_column_letter(c)].width = 20
    for c in range(5, 20):
        ws.column_dimensions[get_column_letter(c)].width = 13

    add_logo(ws, logo_path)
    apply_cell(ws, 1, 2, f"{client_name} — CIP Dashboard",
               font=font_heading(14, True, NAVY), alignment=ALIGN_VCENTER)

    ip_count = len(ip_grouped)
    comp_count = len(comp_grouped)
    div_str = ", ".join(divisions)
    apply_cell(ws, 2, 2,
        f"Run Date: {run_date} | In Progress: {ip_count} jobs | "
        f"Completed ({completed_label}): {comp_count} jobs | Division: {div_str}",
        font=font_body(9, color=HERON))

    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = FILL_OLIVE
    ws.row_dimensions[3].height = 4

    # ── KPIs ──
    if len(ip_grouped) > 0:
        ip_est_rev = ip_grouped['Revenue Estimated'].sum()
        ip_earned = ip_grouped['Earned Revenue'].sum()
        ip_invoiced = ip_grouped['Invoiced Revenue'].sum()
        ip_backlog = ip_grouped['Backlog'].sum()
        ip_act_cost = ip_grouped['Act Cost $'].sum()
        ip_est_cost = ip_grouped['Estimated Cost $'].sum()
        ip_avg_rev_pct = safe_div(ip_earned, ip_est_rev)
        ip_act_gm = 1 - safe_div(ip_act_cost, ip_earned)
        ip_est_gm = 1 - safe_div(ip_est_cost, ip_est_rev)
    else:
        ip_est_rev = ip_earned = ip_invoiced = ip_backlog = ip_act_cost = ip_est_cost = 0
        ip_avg_rev_pct = ip_act_gm = ip_est_gm = 0

    # ── COMMENTED OUT: Portfolio Overview & Completed KPIs (revisit later) ──
    # row = 5
    # ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    # apply_cell(ws, row, 1, "PORTFOLIO OVERVIEW — IN PROGRESS",
    #            font=font_heading(11, True, WHITE_HEX), fill=FILL_NAVY, alignment=ALIGN_CENTER)
    # row += 1
    # for i, (lbl, val, ft) in enumerate([
    #     ("Total Est Revenue", ip_est_rev, 'dollar'), ("Total Earned Revenue", ip_earned, 'dollar'),
    #     ("Total Invoiced", ip_invoiced, 'dollar'), ("Total Backlog", ip_backlog, 'dollar'),
    #     ("Actual Cost to Date", ip_act_cost, 'dollar'),
    # ]):
    #     write_kpi(ws, row, i*2+1, lbl, val, ft)
    #
    # row += 3
    # for i, (lbl, val, ft) in enumerate([
    #     ("Avg Rev % Complete", ip_avg_rev_pct, 'pct'), ("Blended Actual GM%", ip_act_gm, 'pct'),
    #     ("Blended Est GM%", ip_est_gm, 'pct'), ("Active Jobs", ip_count, 'count'),
    # ]):
    #     write_kpi(ws, row, i*2+1, lbl, val, ft)
    #
    # row += 3
    # # Completed KPIs
    # if len(comp_grouped) > 0:
    #     c_est_rev = comp_grouped['Revenue Estimated'].sum()
    #     c_earned = comp_grouped['Earned Revenue'].sum()
    #     c_act_cost = comp_grouped['Act Cost $'].sum()
    #     c_est_cost = comp_grouped['Estimated Cost $'].sum()
    #     c_gm = 1 - safe_div(c_act_cost, c_earned)
    # else:
    #     c_est_rev = c_earned = c_act_cost = c_est_cost = c_gm = 0
    #
    # ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    # apply_cell(ws, row, 1, f"COMPLETED PORTFOLIO (LAST {completed_months} MONTHS)",
    #            font=font_heading(11, True, WHITE_HEX), fill=FILL_NAVY, alignment=ALIGN_CENTER)
    # row += 1
    # for i, (lbl, val, ft) in enumerate([
    #     ("Total Est Revenue", c_est_rev, 'dollar'), ("Total Earned", c_earned, 'dollar'),
    #     ("Actual Cost", c_act_cost, 'dollar'), ("Est Cost", c_est_cost, 'dollar'),
    #     ("Blended GM%", c_gm, 'pct'),
    # ]):
    #     write_kpi(ws, row, i*2+1, lbl, val, ft)
    #
    # row += 4
    # ── END COMMENTED OUT SECTION ──

    row = 5

    # ── Total Backlog by Division (ALL in-process opps, including below the $ threshold) ──
    # Backlog $   = Σ (Estimated Revenue − Earned Revenue)
    # Backlog Hrs = Σ max(Estimated Hours − Actual Hours, 0)  — over-budget jobs don't offset
    bl_src = ip_grouped_all if (ip_grouped_all is not None and len(ip_grouped_all) > 0) else ip_grouped
    if bl_src is not None and len(bl_src) > 0:
        tmp = bl_src.copy()
        tmp['_bl_dollars'] = tmp['Revenue Estimated'].fillna(0) - tmp['Earned Revenue'].fillna(0)
        tmp['_bl_hours'] = (tmp['Labor Hours Estimated'].fillna(0) -
                            tmp['Labor Hours Actual'].fillna(0)).clip(lower=0)
        bl = tmp.groupby('Division').agg(
            _jobs=('Opportunity #', 'count'),
            _bl_dollars=('_bl_dollars', 'sum'),
            _bl_hours=('_bl_hours', 'sum'),
        ).reset_index().sort_values('_bl_dollars', ascending=False)
        bld = [[r['Division'], r['_jobs'], r['_bl_dollars'], r['_bl_hours']]
               for _, r in bl.iterrows()]
        bld.append(['Grand Total', int(bl['_jobs'].sum()),
                    bl['_bl_dollars'].sum(), bl['_bl_hours'].sum()])
        blh = ['Division', 'Open Jobs', 'Backlog $ (Est Rev − Earned)', 'Backlog Hours (Est − Act, floored)']
        blf = ['text', 'text', 'dollar', 'hours']
        end_row = write_dash_table(ws, row,
            "TOTAL BACKLOG BY DIVISION (ALL OPEN JOBS — INCLUDES SUB-THRESHOLD)", blh, bld, blf)
        # Bold the Grand Total row
        gt_r = row + 1 + len(bld)
        for ci in range(1, len(blh) + 1):
            ws.cell(row=gt_r, column=ci).font = font_body(9, bold=True)
            ws.cell(row=gt_r, column=ci).fill = FILL_SUBTOTAL
        row = end_row + 1

    def cf_gm_cells(data, start_row, act_ci, est_ci):
        for ri, rd in enumerate(data):
            r_num = start_row + 2 + ri
            act_v, est_v = rd[act_ci], rd[est_ci]
            if act_v < est_v:
                ws.cell(row=r_num, column=act_ci+1).fill = FILL_RED
            elif act_v > est_v and act_v > 0:
                ws.cell(row=r_num, column=act_ci+1).fill = FILL_GREEN

    # Top 5
    if len(ip_grouped) > 0:
        top5 = ip_grouped.nlargest(5, 'Revenue Estimated')
        t5d = [[r['Opportunity #'], r['Opp Name'], r['Property Name'], r['Revenue Estimated'],
                r['Earned Revenue'], r['Backlog'], r['Act Cost $'], r['Actual GM%'],
                r['Estimated GM%'], r['Branch'], r['Division']] for _, r in top5.iterrows()]
        t5h = ['Opp #', 'Opportunity Name', 'Property', 'Estimated Rev ($)', 'Earned Rev ($)',
               'Backlog', 'Act Cost ($)', 'Act GM%', 'Est GM%', 'Branch', 'Division']
        t5f = ['text','text','text','dollar','dollar','dollar','dollar','pct','pct','text','text']
        end_row = write_dash_table(ws, row, "TOP 5 LARGEST JOBS (BY ESTIMATED REVENUE)", t5h, t5d, t5f)
        cf_gm_cells(t5d, row, 7, 8)
        row = end_row + 1

    # Budget Overruns
    if len(ip_grouped) > 0:
        ov = ip_grouped[(ip_grouped['Act Cost $'] > ip_grouped['Estimated Cost $']) &
                        (ip_grouped['Estimated Cost $'] > 0)].copy()
        ov['Overage $'] = ov['Act Cost $'] - ov['Estimated Cost $']
        ov = ov.sort_values('Overage $', ascending=False)
        ovd = [[r['Opportunity #'], r['Opp Name'], r['Property Name'], r['Act Cost $'],
                r['Estimated Cost $'], r['Overage $'], r['Total Act/Est Cost'],
                r['Actual GM%'], r['Estimated GM%'], r['Branch'], r['Division']]
               for _, r in ov.iterrows()]
        ovh = ['Opp #', 'Opportunity Name', 'Property', 'Act Cost ($)', 'Est Cost ($)',
               'Overage $', 'Cost Ratio', 'Act GM%', 'Est GM%', 'Branch', 'Division']
        ovf = ['text','text','text','dollar','dollar','dollar','pct','pct','pct','text','text']
        end_row = write_dash_table(ws, row, f"BUDGET OVERRUNS ({len(ov)} JOBS)", ovh, ovd, ovf)
        for ri, rd in enumerate(ovd):
            r_num = row + 2 + ri
            if rd[6] > 1: ws.cell(row=r_num, column=7).fill = FILL_RED
        cf_gm_cells(ovd, row, 7, 8)
        row = end_row + 1

    # Cash Flow Risks
    if len(ip_grouped) > 0:
        cfr = ip_grouped[(ip_grouped['Rev % Completed'] > 0.10) &
                         ((ip_grouped['Rev % Completed'] - ip_grouped['Invoice %']) > 0.10)].copy()
        cfr['Gap %'] = cfr['Rev % Completed'] - cfr['Invoice %']
        cfr = cfr.sort_values('Gap %', ascending=False)
        cfd = [[r['Opportunity #'], r['Opp Name'], r['Property Name'], r['Invoice Type'],
                r['Revenue Estimated'], r['Invoiced Revenue'], r['Invoice %'],
                r['Gap %'], r['Earned Revenue'], r['Branch'], r['Division']]
               for _, r in cfr.iterrows()]
        cfh = ['Opp #', 'Opportunity Name', 'Property', 'Invoice Type', 'Estimated Rev ($)',
               'Invoiced Rev ($)', 'Invoice %', 'Gap %', 'Earned Rev ($)', 'Branch', 'Division']
        cff = ['text','text','text','text','dollar','dollar','pct','pct','dollar','text','text']
        end_row = write_dash_table(ws, row, f"CASH FLOW RISKS — UNDER-INVOICED ({len(cfr)} JOBS)",
                                   cfh, cfd, cff)
        row = end_row + 1

    # Labor Concerns
    if len(ip_grouped) > 0:
        lb = ip_grouped[(ip_grouped['Labor Hrs Act/Est'] > 1) &
                        (ip_grouped['Labor Hours Estimated'] > 0)].copy()
        lb['Overage Hrs'] = lb['Labor Hours Actual'] - lb['Labor Hours Estimated']
        lb = lb.sort_values('Labor Hrs Act/Est', ascending=False)
        lbd = [[r['Opportunity #'], r['Opp Name'], r['Property Name'], r['Revenue Estimated'],
                r['Labor Hours Actual'], r['Labor Hours Estimated'], r['Labor Hrs Act/Est'],
                r['Overage Hrs'], r['Labor Cost Actual'], r['Labor Cost Estimated.1'],
                r['Branch'], r['Division']] for _, r in lb.iterrows()]
        lbh = ['Opp #', 'Opportunity Name', 'Property', 'Estimated Rev ($)', 'Act Hours',
               'Est Hours', 'Labor Efficiency (%)', 'Overage Hrs', 'Act Labor ($)',
               'Est Labor ($)', 'Branch', 'Division']
        lbf = ['text','text','text','dollar','hours','hours','pct','hours','dollar','dollar','text','text']
        end_row = write_dash_table(ws, row, f"LABOR CONCERNS — ACT/EST HRS > 100% ({len(lb)} JOBS)",
                                   lbh, lbd, lbf)
        for ri, rd in enumerate(lbd):
            r_num = row + 2 + ri
            if rd[6] > 1: ws.cell(row=r_num, column=7).fill = FILL_RED
        row = end_row + 1

    return ip_earned  # for reconciliation


# ── Column spec type: (header, source_col, width, fmt_type) ──

IP_COLS = [
    ('Opp #', 'Opportunity #', 15, 'text'), ('Property Name', 'Property Name', 25, 'text'),
    ('Opp Name', 'Opp Name', 25, 'text'), ('Company Name', 'Company Name', 15, 'text'),
    ('Sales Rep', 'Sales Rep', 13, 'text'), ('Ops Mgr', 'Ops Mgr', 13, 'text'),
    ('Start Date', 'Start Date Final', 13, 'date'),
    ('Branch', 'Branch', 13, 'text'), ('Division', 'Division', 13, 'text'),
    ('Invoice Type', 'Invoice Type', 13, 'text'),
    ('Earned Revenue', 'Earned Revenue', 13, 'dollar'), ('Estimated Revenue', 'Revenue Estimated', 13, 'dollar'),
    ('Invoiced Revenue', 'Invoiced Revenue', 13, 'dollar'),
    ('Act Labor Hours', 'Labor Hours Actual', 13, 'hours'), ('Est Labor Hours', 'Labor Hours Estimated', 13, 'hours'),
    ('Future Sched Hours', 'Future Scheduled Hours', 13, 'hours'),
    ('Act Labor Cost', 'Labor Cost Actual', 13, 'dollar'), ('Est Labor Cost', 'Labor Cost Estimated.1', 13, 'dollar'),
    ('Act Material Cost', 'Material Cost Actual', 13, 'dollar'), ('Est Material Cost', 'Material Cost Estimated.1', 13, 'dollar'),
    ('Act Sub Cost', 'Sub Cost Actual', 13, 'dollar'), ('Est Sub Cost', 'Sub Cost Estimated.1', 13, 'dollar'),
    ('Act Equipment Cost', 'Equipment Cost Actual', 13, 'dollar'), ('Est Equipment Cost', 'Equipment Cost Estimated.1', 13, 'dollar'),
    ('Act Other Cost', 'Other Cost Actual', 13, 'dollar'), ('Est Other Cost', 'Other Cost Estimated.1', 13, 'dollar'),
    ('Act Cost $', 'Act Cost $', 13, 'dollar'), ('Estimated Cost $', 'Estimated Cost $', 13, 'dollar'),
    ('Actual GM%', 'Actual GM%', 13, 'pct'), ('Estimated GM%', 'Estimated GM%', 13, 'pct'),
    ('Rev % Completed', 'Rev % Completed', 13, 'pct'), ('Invoice %', 'Invoice %', 13, 'pct'),
    ('Total Act/Est Cost', 'Total Act/Est Cost', 13, 'pct'), ('Labor Hrs Act/Est', 'Labor Hrs Act/Est', 13, 'pct'),
    ('Labor $ Act/Est', 'Labor $ Act/Est', 13, 'pct'), ('Materials Act/Est', 'Materials Act/Est', 13, 'pct'),
    ('Sub Act/Est', 'Sub Act/Est', 13, 'pct'), ('Equip Act/Est', 'Equip Act/Est', 13, 'pct'),
    ('Other Act/Est', 'Other Act/Est', 13, 'pct'),
    ('Act+Sched Hrs', 'Actual+Sched Hrs', 13, 'hours'), ('Potential Hr Overage', 'Potential Hr Overage', 13, 'hours'),
    ('Note', 'Note', 30, 'text'),
    ('Hours Remain', 'Hours Remain', 13, 'hours'), ('Subs Remain', 'Subs Remain', 13, 'dollar'),
    ('Mat Remain', 'Mat Remain', 13, 'dollar'), ('Year-Month', 'Year-Month', 13, 'text'),
    ('Backlog', 'Backlog', 13, 'dollar'),
]

BP_COLS = [
    ('Property Name', 'Property Name', 25, 'text'), ('Company Name', 'Company Name', 15, 'text'),
    ('Division', 'Division', 13, 'text'), ('Branch', 'Branch', 13, 'text'),
    ('Earned Revenue', 'Earned Revenue', 13, 'dollar'), ('Estimated Revenue', 'Revenue Estimated', 13, 'dollar'),
    ('Invoiced Revenue', 'Invoiced Revenue', 13, 'dollar'),
    ('Act Labor Hours', 'Labor Hours Actual', 13, 'hours'), ('Est Labor Hours', 'Labor Hours Estimated', 13, 'hours'),
    ('Future Sched Hours', 'Future Scheduled Hours', 13, 'hours'),
    ('Act Labor Cost', 'Labor Cost Actual', 13, 'dollar'), ('Est Labor Cost', 'Labor Cost Estimated.1', 13, 'dollar'),
    ('Act Material Cost', 'Material Cost Actual', 13, 'dollar'), ('Est Material Cost', 'Material Cost Estimated.1', 13, 'dollar'),
    ('Act Sub Cost', 'Sub Cost Actual', 13, 'dollar'), ('Est Sub Cost', 'Sub Cost Estimated.1', 13, 'dollar'),
    ('Act Equipment Cost', 'Equipment Cost Actual', 13, 'dollar'), ('Est Equipment Cost', 'Equipment Cost Estimated.1', 13, 'dollar'),
    ('Act Other Cost', 'Other Cost Actual', 13, 'dollar'), ('Est Other Cost', 'Other Cost Estimated.1', 13, 'dollar'),
    ('Act Cost $', 'Act Cost $', 13, 'dollar'), ('Estimated Cost $', 'Estimated Cost $', 13, 'dollar'),
    ('Actual GM%', 'Actual GM%', 13, 'pct'), ('Estimated GM%', 'Estimated GM%', 13, 'pct'),
    ('Rev % Completed', 'Rev % Completed', 13, 'pct'), ('Invoice %', 'Invoice %', 13, 'pct'),
    ('Total Act/Est Cost', 'Total Act/Est Cost', 13, 'pct'), ('Labor Hrs Act/Est', 'Labor Hrs Act/Est', 13, 'pct'),
    ('Labor $ Act/Est', 'Labor $ Act/Est', 13, 'pct'), ('Materials Act/Est', 'Materials Act/Est', 13, 'pct'),
    ('Sub Act/Est', 'Sub Act/Est', 13, 'pct'), ('Equip Act/Est', 'Equip Act/Est', 13, 'pct'),
    ('Other Act/Est', 'Other Act/Est', 13, 'pct'),
    ('Act+Sched Hrs', 'Actual+Sched Hrs', 13, 'hours'), ('Potential Hr Overage', 'Potential Hr Overage', 13, 'hours'),
    ('Hours Remain', 'Hours Remain', 13, 'hours'), ('Subs Remain', 'Subs Remain', 13, 'dollar'),
    ('Mat Remain', 'Mat Remain', 13, 'dollar'), ('Year-Month', 'Year-Month', 13, 'text'),
]

COMP_COLS = [
    ('Opp #', 'Opportunity #', 15, 'text'), ('Property Name', 'Property Name', 25, 'text'),
    ('Opp Name', 'Opp Name', 25, 'text'), ('Company Name', 'Company Name', 15, 'text'),
    ('Sales Rep', 'Sales Rep', 13, 'text'), ('Ops Mgr', 'Ops Mgr', 13, 'text'),
    ('Start Date', 'Start Date Final', 13, 'date'),
    ('Branch', 'Branch', 13, 'text'), ('Division', 'Division', 13, 'text'),
    ('Invoice Type', 'Invoice Type', 13, 'text'),
    ('Complete Date', 'Oppty Complete Date', 13, 'date'),
    ('Earned Revenue', 'Earned Revenue', 13, 'dollar'), ('Estimated Revenue', 'Revenue Estimated', 13, 'dollar'),
    ('Invoiced Revenue', 'Invoiced Revenue', 13, 'dollar'),
    ('Act Labor Hours', 'Labor Hours Actual', 13, 'hours'), ('Est Labor Hours', 'Labor Hours Estimated', 13, 'hours'),
    ('Future Sched Hours', 'Future Scheduled Hours', 13, 'hours'),
    ('Act Labor Cost', 'Labor Cost Actual', 13, 'dollar'), ('Est Labor Cost', 'Labor Cost Estimated.1', 13, 'dollar'),
    ('Act Material Cost', 'Material Cost Actual', 13, 'dollar'), ('Est Material Cost', 'Material Cost Estimated.1', 13, 'dollar'),
    ('Act Sub Cost', 'Sub Cost Actual', 13, 'dollar'), ('Est Sub Cost', 'Sub Cost Estimated.1', 13, 'dollar'),
    ('Act Equipment Cost', 'Equipment Cost Actual', 13, 'dollar'), ('Est Equipment Cost', 'Equipment Cost Estimated.1', 13, 'dollar'),
    ('Act Other Cost', 'Other Cost Actual', 13, 'dollar'), ('Est Other Cost', 'Other Cost Estimated.1', 13, 'dollar'),
    ('Act Cost $', 'Act Cost $', 13, 'dollar'), ('Estimated Cost $', 'Estimated Cost $', 13, 'dollar'),
    ('Actual GM%', 'Actual GM%', 13, 'pct'), ('Estimated GM%', 'Estimated GM%', 13, 'pct'),
    ('Rev % Completed', 'Rev % Completed', 13, 'pct'), ('Invoice %', 'Invoice %', 13, 'pct'),
    ('Total Act/Est Cost', 'Total Act/Est Cost', 13, 'pct'), ('Labor Hrs Act/Est', 'Labor Hrs Act/Est', 13, 'pct'),
    ('Labor $ Act/Est', 'Labor $ Act/Est', 13, 'pct'), ('Materials Act/Est', 'Materials Act/Est', 13, 'pct'),
    ('Sub Act/Est', 'Sub Act/Est', 13, 'pct'), ('Equip Act/Est', 'Equip Act/Est', 13, 'pct'),
    ('Other Act/Est', 'Other Act/Est', 13, 'pct'),
    ('Act+Sched Hrs', 'Actual+Sched Hrs', 13, 'hours'), ('Potential Hr Overage', 'Potential Hr Overage', 13, 'hours'),
    ('Note', 'Note', 30, 'text'),
    ('Actual Gross Profit', 'Actual Gross Profit', 13, 'dollar'),
    ('Est Gross Profit', 'Est Gross Profit', 13, 'dollar'),
    ('Year-Month', 'Year-Month', 13, 'text'),
    ('Complete Year', 'Complete Year', 13, 'text'), ('Complete Quarter', 'Complete Quarter', 13, 'text'),
]


def build_data_tab(wb, tab_name, tab_cols, df, client_name, logo_path, run_date,
                   invoice_flag_rule="lag", invoice_flag_gap=0.10, cost_pace_threshold=0.0):
    is_inprocess = tab_name in ("In Process", "In Process By Property")
    ws = wb.create_sheet(tab_name)
    ws.sheet_properties.tabColor = TAB_COLORS.get(tab_name, HERON)
    ws.sheet_view.showGridLines = True
    ncols = len(tab_cols)
    last_col = get_column_letter(ncols)

    add_logo(ws, logo_path)

    # Row 1: Title
    ws.merge_cells(f'B1:{last_col}1')
    apply_cell(ws, 1, 2, f"{client_name} — {tab_name}",
               font=font_heading(12, True, NAVY), alignment=ALIGN_VCENTER_NOWRAP)

    # Row 2: Run date
    ws.merge_cells(f'B2:{last_col}2')
    apply_cell(ws, 2, 2, f"Run Date: {run_date}",
               font=font_body(9, color=HERON), alignment=ALIGN_VCENTER_NOWRAP)

    # Row 3: Grand Total (formulas — written after data rows so we know the range)
    # Placeholder — filled below after data rows are written

    # Row 4: CF notes (+ editable cost-pace threshold cell on In Process tabs)
    PACE_COLS = ['Total Act/Est Cost', 'Labor Hrs Act/Est', 'Labor $ Act/Est',
                 'Materials Act/Est', 'Sub Act/Est', 'Equip Act/Est', 'Other Act/Est']
    col_idx_pre = {hdr: i + 1 for i, (hdr, _, _, _) in enumerate(tab_cols)}
    pace_cell_ref = None
    if is_inprocess:
        # Put the editable threshold in the row-4 cell of Total Act/Est Cost (or first pace col present)
        anchor_hdr = next((h for h in PACE_COLS if h in col_idx_pre), None)
        cf_note_map = {
            'Actual GM%': 'Green=Act>Est; Red=Act<Est (blank if not started)',
            'Invoice %': ('Yellow if >10% below Rev%' if invoice_flag_rule == 'lag'
                          else 'Yellow if > Earned'),
        }
        if anchor_hdr:
            pace_ci = col_idx_pre[anchor_hdr]
            pace_cell_ref = f'${get_column_letter(pace_ci)}$4'
            apply_cell(ws, 4, pace_ci, cost_pace_threshold, font=font_body(9, bold=True, color=NAVY),
                       fill=FILL_YELLOW, alignment=ALIGN_PCT, number_format=FMT_PCT, border=THIN_BORDER)
            # Label in the cell to the left, if free
            if pace_ci - 1 >= 1:
                apply_cell(ws, 4, pace_ci - 1, 'Pace threshold →',
                           font=FONT_CF_NOTE, alignment=ALIGN_RIGHT)
        for ci, (hdr, _, _, _) in enumerate(tab_cols, 1):
            if hdr in PACE_COLS:
                continue  # threshold cell / driven by CF rule
            note = cf_note_map.get(hdr)
            if note:
                apply_cell(ws, 4, ci, note, font=FONT_CF_NOTE, alignment=ALIGN_PCT)
    else:
        cf_note_map = {
            'Actual GM%': 'Green=Act>Est; Red=Act<Est (blank if not started)',
        }
        for ci, (hdr, _, _, _) in enumerate(tab_cols, 1):
            note = cf_note_map.get(hdr) or ('Red if >100%' if 'Act/Est' in hdr else None)
            if note:
                apply_cell(ws, 4, ci, note, font=FONT_CF_NOTE, alignment=ALIGN_PCT)

    # Row 5: Headers
    for ci, (hdr, _, w, _) in enumerate(tab_cols, 1):
        apply_cell(ws, 5, ci, hdr, font=font_body(10, bold=True, color=WHITE_HEX),
                   fill=FILL_HEADER, alignment=ALIGN_CENTER, border=THIN_BORDER)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.column_dimensions['A'].width = max(tab_cols[0][2], 15)

    if len(df) == 0:
        apply_cell(ws, 3, 1, "Grand Total", font=font_body(10, bold=True), fill=FILL_SUBTOTAL)
        for ci in range(1, ncols + 1):
            cell = ws.cell(row=3, column=ci)
            cell.fill = FILL_SUBTOTAL
            cell.font = font_body(10, bold=True)
            cell.border = THIN_BORDER
        apply_cell(ws, 6, 1, "No data", font=font_body(10))
        return ws

    # Data rows
    for ri, (_, row_data) in enumerate(df.iterrows()):
        r = 6 + ri
        fill = FILL_ZEBRA if ri % 2 == 0 else FILL_WHITE
        for ci, (hdr, src, _, fmt) in enumerate(tab_cols, 1):
            val = row_data.get(src, '') if src in df.columns else ''
            # Potential Hr Overage is NaN when not over budget — leave blank, don't show 0.
            if hdr == 'Potential Hr Overage' and (not isinstance(val, str)) and pd.isna(val):
                apply_cell(ws, r, ci, '', font=font_body(10), fill=fill,
                           alignment=ALIGN_RIGHT, border=THIN_BORDER)
                continue
            if pd.isna(val) if not isinstance(val, str) else False:
                val = 0 if fmt in ('dollar', 'pct', 'hours') else ''

            if fmt == 'dollar':
                nf, al = FMT_DOLLAR, ALIGN_RIGHT
            elif fmt == 'pct':
                nf, al = FMT_PCT, ALIGN_PCT
            elif fmt == 'hours':
                nf, al = FMT_HOURS, ALIGN_RIGHT
            elif fmt == 'date':
                if isinstance(val, pd.Timestamp):
                    val = val.to_pydatetime()
                nf, al = FMT_DATE, ALIGN_CENTER
            else:
                nf, al = None, ALIGN_LEFT

            apply_cell(ws, r, ci, val, font=font_body(10), fill=fill,
                       alignment=al, number_format=nf, border=THIN_BORDER)

    # ── Grand Total in Row 3 with SUBTOTAL formulas ──
    last_data_row = 5 + len(df)  # row 6 is first data row, last is 6 + len(df) - 1
    apply_cell(ws, 3, 1, "Grand Total", font=font_body(10, bold=True), fill=FILL_SUBTOTAL,
               border=THIN_BORDER)

    # Build column-letter lookup for formula references
    col_idx = {hdr: i+1 for i, (hdr, _, _, _) in enumerate(tab_cols)}

    # Map percentage/calculated headers to formulas referencing row 3 subtotal cells.
    # These reference the already-calculated SUBTOTAL cells in row 3 rather than
    # recomputing SUBTOTAL ranges, keeping the Grand Total row self-consistent.
    # Format: header -> (formula_type, numerator_col_header, denominator_col_header)
    #   'gm'    → =IFERROR((revenue - cost) / revenue, "-")
    #   'ratio' → =IFERROR(num / den, "-")
    pct_formula_map = {
        'Actual GM%': ('gm', 'Earned Revenue', 'Act Cost $'),
        'Estimated GM%': ('gm', 'Estimated Revenue', 'Estimated Cost $'),
        'Rev % Completed': ('ratio', 'Earned Revenue', 'Estimated Revenue'),
        'Invoice %': ('ratio', 'Invoiced Revenue', 'Earned Revenue'),
        'Total Act/Est Cost': ('ratio', 'Act Cost $', 'Estimated Cost $'),
        'Labor Hrs Act/Est': ('ratio', 'Act Labor Hours', 'Est Labor Hours'),
        'Labor $ Act/Est': ('ratio', 'Act Labor Cost', 'Est Labor Cost'),
        'Materials Act/Est': ('ratio', 'Act Material Cost', 'Est Material Cost'),
        'Sub Act/Est': ('ratio', 'Act Sub Cost', 'Est Sub Cost'),
        'Equip Act/Est': ('ratio', 'Act Equipment Cost', 'Est Equipment Cost'),
        'Other Act/Est': ('ratio', 'Act Other Cost', 'Est Other Cost'),
    }

    # Map calculated hours headers to formulas referencing row 3 subtotal cells.
    # These override the default SUBTOTAL(109,range) for these specific columns.
    # Format: header -> (operand1_col_header, operator, operand2_col_header)
    calc_hours_map = {
        'Act+Sched Hrs': ('Act Labor Hours', '+', 'Future Sched Hours'),
        'Potential Hr Overage': ('Act+Sched Hrs', '-', 'Est Labor Hours'),
    }

    for ci, (hdr, src, _, fmt) in enumerate(tab_cols, 1):
        cell = ws.cell(row=3, column=ci)
        cell.fill = FILL_SUBTOTAL
        cell.font = font_body(10, bold=True)
        cell.border = THIN_BORDER
        col_letter = get_column_letter(ci)

        if hdr in calc_hours_map:
            # Calculated hours columns: reference row 3 subtotal cells
            op1_hdr, operator, op2_hdr = calc_hours_map[hdr]
            op1_ci = col_idx.get(op1_hdr)
            op2_ci = col_idx.get(op2_hdr)
            if op1_ci and op2_ci:
                op1_letter = get_column_letter(op1_ci)
                op2_letter = get_column_letter(op2_ci)
                if hdr == 'Potential Hr Overage':
                    # Only show a positive overage; blank otherwise (mirrors per-row rule)
                    cell.value = (f'=IF({op1_letter}3{operator}{op2_letter}3>0,'
                                  f'{op1_letter}3{operator}{op2_letter}3,"")')
                else:
                    cell.value = f'={op1_letter}3{operator}{op2_letter}3'
            cell.number_format = FMT_HOURS
            cell.alignment = ALIGN_RIGHT
        elif fmt in ('dollar', 'hours') and src in df.columns:
            # SUBTOTAL(109, range) = SUM of visible cells only
            cell.value = f'=SUBTOTAL(109,{col_letter}6:{col_letter}{last_data_row})'
            if fmt == 'dollar':
                cell.number_format = FMT_DOLLAR
                cell.alignment = ALIGN_RIGHT
            else:
                cell.number_format = FMT_HOURS
                cell.alignment = ALIGN_RIGHT
        elif fmt == 'pct' and hdr in pct_formula_map:
            ptype, num_hdr, den_hdr = pct_formula_map[hdr]
            num_ci = col_idx.get(num_hdr)
            den_ci = col_idx.get(den_hdr)
            if num_ci and den_ci:
                num_letter = get_column_letter(num_ci)
                den_letter = get_column_letter(den_ci)
                if ptype == 'gm':
                    # GM% = (revenue - cost) / revenue
                    cell.value = f'=IFERROR(({num_letter}3-{den_letter}3)/{num_letter}3,"-")'
                else:
                    # Ratio = num/den
                    cell.value = f'=IFERROR({num_letter}3/{den_letter}3,"-")'
            cell.number_format = FMT_PCT
            cell.alignment = ALIGN_PCT

    # Conditional formatting (cell-level)
    act_gm_ci = col_idx.get('Actual GM%')
    est_gm_ci = col_idx.get('Estimated GM%')
    rev_pct_ci = col_idx.get('Rev % Completed')
    inv_pct_ci = col_idx.get('Invoice %')
    ratio_cis = [ci for hdr, ci in col_idx.items() if 'Act/Est' in hdr]

    for ri in range(len(df)):
        r = 6 + ri
        rd = df.iloc[ri]
        earned = rd.get('Earned Revenue', 0)
        not_started = (earned is None or earned == 0 or
                       (isinstance(earned, float) and np.isnan(earned)))

        # GM%: no highlight when the job hasn't started (0% is expected, not a problem)
        if act_gm_ci and est_gm_ci and not not_started:
            av, ev = rd.get('Actual GM%', 0), rd.get('Estimated GM%', 0)
            if av < ev:
                ws.cell(row=r, column=act_gm_ci).fill = FILL_RED
            elif av > ev and av > 0:
                ws.cell(row=r, column=act_gm_ci).fill = FILL_GREEN

        # Orange Rev % Completed ≥ 100% — In Process only (completed jobs are all ~100%)
        if is_inprocess and rev_pct_ci and rd.get('Rev % Completed', 0) >= 1:
            ws.cell(row=r, column=rev_pct_ci).fill = FILL_ORANGE

        # Invoice % highlight
        if inv_pct_ci:
            inv = rd.get('Invoice %', 0)
            revc = rd.get('Rev % Completed', 0)
            flag = False
            if is_inprocess:
                if invoice_flag_rule == 'over':
                    flag = inv > revc
                else:  # lag: more than gap below Rev % Completed
                    flag = (revc - inv) > invoice_flag_gap
            else:
                flag = inv < revc  # Completed: simple lag
            if flag:
                ws.cell(row=r, column=inv_pct_ci).fill = FILL_YELLOW

        # Static Act/Est >100% red — Completed tab only (In Process uses dynamic CF below)
        if not is_inprocess:
            for ci in ratio_cis:
                src = tab_cols[ci-1][1]
                if src in df.columns and rd.get(src, 0) > 1:
                    ws.cell(row=r, column=ci).fill = FILL_RED

        # Blue "not started" Property highlight — In Process only
        if is_inprocess and not_started:
            prop_ci = col_idx.get('Property Name')
            if prop_ci:
                ws.cell(row=r, column=prop_ci).fill = FILL_NOT_STARTED

    # Dynamic, editable Act/Est pace CF on In Process tabs.
    # A cell turns red when (Act/Est ratio − Rev % Completed) exceeds the threshold in
    # the row-4 cell (pace_cell_ref). Editing that cell re-evaluates the formatting live.
    if is_inprocess and pace_cell_ref and rev_pct_ci and len(df) > 0:
        rev_letter = get_column_letter(rev_pct_ci)
        for ci in ratio_cis:
            col_letter = get_column_letter(ci)
            rng = f'{col_letter}6:{col_letter}{last_data_row}'
            formula = f'($ {col_letter}6-$ {rev_letter}6)>{pace_cell_ref}'.replace(' ', '')
            ws.conditional_formatting.add(
                rng, FormulaRule(formula=[formula], fill=FILL_RED, stopIfTrue=False))

    # Freeze + autofilter (autofilter over header + data rows only, not the grand total row)
    # Tab-specific freeze: freeze columns through the key identifier column + rows 1-5
    freeze_col_map = {
        'In Process': 'Opp Name',           # Freeze through col C (Opp #, Property, Opp Name)
        'In Process By Property': 'Property Name',  # Freeze through col A (Property Name)
        'Completed': 'Property Name',        # Freeze through col B (Opp #, Property Name)
    }
    freeze_hdr = freeze_col_map.get(tab_name)
    if freeze_hdr and freeze_hdr in col_idx:
        freeze_col_num = col_idx[freeze_hdr] + 1  # +1 to freeze THROUGH that column
        ws.freeze_panes = f'{get_column_letter(freeze_col_num)}6'
    else:
        ws.freeze_panes = 'A6'
    ws.auto_filter.ref = f"A5:{last_col}{last_data_row}"

    # Column grouping — collapse dimension columns (Company … Invoice Type)
    grp1_start = col_idx.get('Company Name')
    grp1_end = col_idx.get('Invoice Type')
    if grp1_start and grp1_end and grp1_end >= grp1_start:
        for c in range(grp1_start, grp1_end + 1):
            ws.column_dimensions[get_column_letter(c)].outlineLevel = 1
            ws.column_dimensions[get_column_letter(c)].hidden = True

    grp2_start = col_idx.get('Act Labor Hours')
    grp2_end = col_idx.get('Est Other Cost')
    if grp2_start and grp2_end and grp2_end >= grp2_start:
        for c in range(grp2_start, grp2_end + 1):
            ws.column_dimensions[get_column_letter(c)].outlineLevel = 2
            ws.column_dimensions[get_column_letter(c)].hidden = True

    return ws


OVERVIEW_HEADERS = [
    '', 'Count', 'Rev', '% of Total Rev', 'Rev per Job',
    'Actual Gross Profit', 'Actual GM', 'Est Gross Profit', 'Est GM',
    'Act Labor Cost', 'Est Labor Cost', 'Act/Est Labor %',
    'Act Sub Cost', 'Est Sub Cost', 'Act/Est Sub %',
    'Act Materials Cost', 'Est Materials Cost', 'Act/Est Materials %',
    'Total Cost', 'Non Sub Costs', 'Non Sub Rev', 'Non Sub GM',
]

OVERVIEW_FMTS = [
    'text', 'count', 'dollar', 'pct', 'dollar', 'dollar', 'pct', 'dollar', 'pct',
    'dollar', 'dollar', 'pct', 'dollar', 'dollar', 'pct',
    'dollar', 'dollar', 'pct', 'dollar', 'dollar', 'dollar', 'pct',
]


def compute_bucket_row(subset, sub_margin=0.281):
    """
    Compute one row for the Complete Overview bucket tables.

    Non Sub Rev backs out the implied revenue attributable to sub work using the
    expected sub margin, isolating margin performance on self-performed work:
        implied_sub_rev = Act Sub Cost / (1 - sub_margin)
        Non Sub Rev    = Earned Revenue - implied_sub_rev
        Non Sub GM     = (Non Sub Rev - Non Sub Costs) / Non Sub Rev
    """
    count = len(subset)
    if count == 0:
        return [0]*21
    rev = subset['Earned Revenue'].sum()
    act_cost = subset['Act Cost $'].sum()
    est_cost = subset['Estimated Cost $'].sum()
    est_rev = subset['Revenue Estimated'].sum()
    act_labor = subset['Labor Cost Actual'].sum()
    est_labor = subset['Labor Cost Estimated.1'].sum()
    act_sub = subset['Sub Cost Actual'].sum()
    est_sub = subset['Sub Cost Estimated.1'].sum()
    act_mat = subset['Material Cost Actual'].sum()
    est_mat = subset['Material Cost Estimated.1'].sum()
    non_sub_costs = act_cost - act_sub
    implied_sub_rev = safe_div(act_sub, 1 - sub_margin)
    non_sub_rev = rev - implied_sub_rev
    return [
        count, rev, 0, safe_div(rev, count),  # pct_rev filled later
        rev - act_cost, safe_div(rev - act_cost, rev),
        est_rev - est_cost, safe_div(est_rev - est_cost, est_rev),
        act_labor, est_labor, safe_div(act_labor, est_labor),
        act_sub, est_sub, safe_div(act_sub, est_sub),
        act_mat, est_mat, safe_div(act_mat, est_mat),
        act_cost, non_sub_costs, non_sub_rev,
        safe_div(non_sub_rev - non_sub_costs, non_sub_rev),
    ]


def build_complete_overview(wb, client_name, logo_path, run_date, comp_grouped,
                            gm_buckets, rev_buckets, sub_margin=0.281):
    ws = wb.create_sheet("Complete Overview")
    ws.sheet_properties.tabColor = TAB_COLORS["Complete Overview"]
    ws.sheet_view.showGridLines = False

    add_logo(ws, logo_path)
    ws.column_dimensions['A'].width = 16
    for c in range(2, 23):
        ws.column_dimensions[get_column_letter(c)].width = 14

    apply_cell(ws, 1, 2, f"{client_name} — Complete Overview",
               font=font_heading(12, True, NAVY), alignment=ALIGN_VCENTER_NOWRAP)
    ws.merge_cells('B2:V2')
    apply_cell(ws, 2, 2, f"Run Date: {run_date}", font=font_body(9, color=HERON))

    total_rev = comp_grouped['Earned Revenue'].sum() if len(comp_grouped) > 0 else 0

    def write_table(start_row, title, buckets, value_col):
        ncols = 22
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=ncols)
        apply_cell(ws, start_row, 1, title, font=font_heading(11, True, WHITE_HEX),
                   fill=FILL_NAVY, alignment=ALIGN_CENTER)

        hr = start_row + 1
        for ci, h in enumerate(OVERVIEW_HEADERS, 1):
            apply_cell(ws, hr, ci, h, font=font_body(9, bold=True, color=WHITE_HEX),
                       fill=FILL_HERON, alignment=ALIGN_CENTER, border=THIN_BORDER)

        for bi, (label, lo, hi) in enumerate(buckets):
            r = hr + 1 + bi
            fill = FILL_ZEBRA if bi % 2 == 0 else FILL_WHITE
            if len(comp_grouped) > 0:
                if value_col == 'Actual GM%':
                    mask = (comp_grouped['Actual GM%'] >= lo) & (comp_grouped['Actual GM%'] < hi) \
                        if hi != float('inf') else (comp_grouped['Actual GM%'] >= lo)
                else:
                    mask = (comp_grouped['Earned Revenue'] >= lo) & (comp_grouped['Earned Revenue'] < hi) \
                        if hi != float('inf') else (comp_grouped['Earned Revenue'] >= lo)
                subset = comp_grouped[mask]
            else:
                subset = pd.DataFrame()

            vals = compute_bucket_row(subset, sub_margin)
            vals[2] = safe_div(vals[1], total_rev)  # pct of total rev

            all_vals = [label] + vals
            for ci, (val, fmt) in enumerate(zip(all_vals, OVERVIEW_FMTS), 1):
                if fmt == 'dollar': nf, al = FMT_DOLLAR, ALIGN_RIGHT
                elif fmt == 'pct': nf, al = FMT_PCT, ALIGN_PCT
                elif fmt == 'count': nf, al = FMT_COUNT, ALIGN_CENTER
                else: nf, al = None, ALIGN_CENTER
                apply_cell(ws, r, ci, val, font=font_body(9), fill=fill,
                           alignment=al, number_format=nf, border=THIN_BORDER)

        # Grand Total
        gt_row = hr + 1 + len(buckets)
        if len(comp_grouped) > 0:
            gt_vals = compute_bucket_row(comp_grouped, sub_margin)
            gt_vals[2] = 1.0
        else:
            gt_vals = [0]*21

        gt_all = ['Grand Total'] + gt_vals
        for ci, (val, fmt) in enumerate(zip(gt_all, OVERVIEW_FMTS), 1):
            if fmt == 'dollar': nf, al = FMT_DOLLAR, ALIGN_RIGHT
            elif fmt == 'pct': nf, al = FMT_PCT, ALIGN_PCT
            elif fmt == 'count': nf, al = FMT_COUNT, ALIGN_CENTER
            else: nf, al = None, ALIGN_CENTER
            apply_cell(ws, gt_row, ci, val, font=font_body(9, bold=True),
                       fill=FILL_SUBTOTAL, alignment=al, number_format=nf, border=THIN_BORDER)

        return gt_row + 2

    row = 4
    row = write_table(row, "GM RANGE ANALYSIS", gm_buckets, 'Actual GM%')
    row = write_table(row, "REVENUE RANGE ANALYSIS", rev_buckets, 'Earned Revenue')


def build_completed_dashboard_view(wb, client_name, logo_path, run_date, comp_grouped):
    """Build the Completed Dashboard View tab with static GP waterfall chart and pivot table."""
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import matplotlib.ticker as mticker
    import tempfile

    ws = wb.create_sheet("Completed Dashboard View")
    ws.sheet_properties.tabColor = TAB_COLORS["Completed Dashboard View"]
    ws.sheet_view.showGridLines = False

    add_logo(ws, logo_path)
    ws.column_dimensions['A'].width = 22
    for c in range(2, 7):
        ws.column_dimensions[get_column_letter(c)].width = 20

    # Row 1: Title
    ws.merge_cells('B1:F1')
    apply_cell(ws, 1, 2, f"{client_name} — Completed Gross Profit Analysis",
               font=font_heading(14, True, NAVY), alignment=ALIGN_VCENTER_NOWRAP)

    # Row 2: Run date
    ws.merge_cells('B2:F2')
    apply_cell(ws, 2, 2, f"Run Date: {run_date}",
               font=font_body(9, color=HERON), alignment=ALIGN_VCENTER_NOWRAP)

    # Olive accent bar
    for c in range(1, 7):
        ws.cell(row=3, column=c).fill = FILL_OLIVE
    ws.row_dimensions[3].height = 4

    # ── Compute GP metrics ──
    if len(comp_grouped) > 0:
        df = comp_grouped.copy()
        df['Est GP'] = df['Revenue Estimated'].fillna(0) - df['Estimated Cost $'].fillna(0)
        df['Act GP'] = df['Earned Revenue'].fillna(0) - df['Act Cost $'].fillna(0)
        df['GP Variance'] = df['Act GP'] - df['Est GP']
        df['GP Over'] = df['GP Variance'].clip(lower=0)
        df['GP Under'] = (-df['GP Variance']).clip(lower=0)

        est_gp_total = df['Est GP'].sum()
        over_total = df['GP Over'].sum()
        under_total = df['GP Under'].sum()
        act_gp_total = df['Act GP'].sum()
        net_variance = over_total - under_total
    else:
        est_gp_total = over_total = under_total = act_gp_total = net_variance = 0
        df = pd.DataFrame()

    # ═══════════════════════════════════════════════════════════════════
    # WATERFALL CHART — rendered as static PNG via matplotlib
    # ═══════════════════════════════════════════════════════════════════

    c_navy = f'#{NAVY}'
    c_olive = f'#{OLIVE}'
    c_mango = f'#{MANGO}'
    c_heron = f'#{HERON}'
    c_red = '#C0504D'

    categories = ['Estimated\nGross Profit', 'GP Over\nBudget', 'GP Under\nBudget', 'Actual\nGross Profit']
    bases =      [0,              est_gp_total,  act_gp_total,  0]
    segments =   [est_gp_total,   over_total,    under_total,   act_gp_total]
    bar_colors = [c_navy,         c_olive,       c_red,         c_mango]

    fig, ax = plt.subplots(figsize=(10, 5.5))
    fig.patch.set_facecolor('white')
    ax.set_facecolor('white')

    x = range(len(categories))
    # Invisible base
    ax.bar(x, bases, color='none', edgecolor='none')
    # Visible segments stacked on base
    bars = ax.bar(x, segments, bottom=bases, color=bar_colors, edgecolor='white', linewidth=0.5, width=0.55)

    # Connector lines between bars
    for i in range(len(categories) - 1):
        connector_y = bases[i] + segments[i]
        ax.plot([i + 0.275, i + 0.725], [connector_y, connector_y],
                color=c_heron, linewidth=0.8, linestyle='--', alpha=0.5)

    # Data labels on each bar
    for i, (base, seg) in enumerate(zip(bases, segments)):
        mid = base + seg / 2
        label = f'${seg:,.0f}'
        ax.text(i, mid, label, ha='center', va='center',
                fontsize=10, fontweight='bold', color='white',
                fontfamily='sans-serif')

    # Title
    ax.set_title('Gross Profit Waterfall — Completed Jobs',
                 fontsize=13, fontweight='bold', color=c_navy,
                 fontfamily='serif', pad=14)

    ax.set_xticks(x)
    ax.set_xticklabels(categories, fontsize=10, fontweight='bold', color=c_navy,
                       fontfamily='sans-serif')
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, p: f'${v:,.0f}'))
    ax.tick_params(axis='y', labelsize=8, labelcolor=c_heron)

    # Clean up spines
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.spines['left'].set_color('#D9D9D9')
    ax.spines['bottom'].set_color('#D9D9D9')
    ax.yaxis.grid(True, color='#EBEBEB', linewidth=0.5, linestyle='-')
    ax.set_axisbelow(True)

    plt.tight_layout()

    # Save to temp file
    chart_path = tempfile.mktemp(suffix='.png')
    fig.savefig(chart_path, dpi=150, bbox_inches='tight', facecolor='white')
    plt.close(fig)

    # Insert chart image anchored to cell A4 (top-left corner)
    chart_img = XLImage(chart_path)
    chart_img.width = 720
    chart_img.height = 396
    ws.add_image(chart_img, "A4")

    # ═══════════════════════════════════════════════════════════════════
    # PIVOT TABLE — Branch x Division with Net GP Variance (single col)
    # ═══════════════════════════════════════════════════════════════════

    pivot_start_row = 28

    ws.merge_cells(f'A{pivot_start_row}:E{pivot_start_row}')
    apply_cell(ws, pivot_start_row, 1, "GROSS PROFIT ANALYSIS BY BRANCH & DIVISION",
               font=font_heading(11, True, WHITE_HEX), fill=FILL_NAVY, alignment=ALIGN_CENTER)

    pivot_headers = ['Branch', 'Division', 'Est Gross Profit',
                     'Net GP Variance', 'Actual Gross Profit']
    pivot_fmts = ['text', 'text', 'dollar', 'dollar', 'dollar']
    hr = pivot_start_row + 1
    for ci, h in enumerate(pivot_headers, 1):
        apply_cell(ws, hr, ci, h, font=font_body(10, bold=True, color=WHITE_HEX),
                   fill=FILL_HERON, alignment=ALIGN_CENTER, border=THIN_BORDER)

    if len(df) > 0:
        pivot = df.groupby(['Branch', 'Division']).agg({
            'Est GP': 'sum', 'GP Variance': 'sum', 'Act GP': 'sum',
        }).reset_index()
        pivot = pivot.sort_values(['Branch', 'Division']).reset_index(drop=True)
    else:
        pivot = pd.DataFrame(columns=['Branch', 'Division', 'Est GP', 'GP Variance', 'Act GP'])

    for ri, (_, prow) in enumerate(pivot.iterrows()):
        r = hr + 1 + ri
        fill = FILL_ZEBRA if ri % 2 == 0 else FILL_WHITE
        row_vals = [prow['Branch'], prow['Division'], prow['Est GP'],
                    prow['GP Variance'], prow['Act GP']]
        for ci, (val, fmt) in enumerate(zip(row_vals, pivot_fmts), 1):
            if fmt == 'dollar':
                nf, al = FMT_DOLLAR, ALIGN_RIGHT
            else:
                nf, al = None, ALIGN_LEFT
            cell = apply_cell(ws, r, ci, val, font=font_body(10), fill=fill,
                              alignment=al, number_format=nf, border=THIN_BORDER)
            if ci == 4 and isinstance(val, (int, float)):
                if val < 0:
                    cell.font = font_body(10, color='C0504D')
                elif val > 0:
                    cell.font = font_body(10, color=OLIVE)

    gt_row = hr + 1 + len(pivot)
    apply_cell(ws, gt_row, 1, "Grand Total", font=font_body(10, bold=True),
               fill=FILL_SUBTOTAL, border=THIN_BORDER)
    apply_cell(ws, gt_row, 2, "", font=font_body(10, bold=True),
               fill=FILL_SUBTOTAL, border=THIN_BORDER)
    gt_vals = [est_gp_total, net_variance, act_gp_total]
    for ci, val in enumerate(gt_vals, 3):
        cell = apply_cell(ws, gt_row, ci, val, font=font_body(10, bold=True),
                          fill=FILL_SUBTOTAL, alignment=ALIGN_RIGHT,
                          number_format=FMT_DOLLAR, border=THIN_BORDER)
        if ci == 4:
            color = OLIVE if val >= 0 else 'C0504D'
            cell.font = font_body(10, bold=True, color=color)


    # Note: chart_path temp file must persist until wb.save() is called.
    # It will be cleaned up by the OS temp file mechanism.


def build_table_tab(wb, raw_df):
    """Create a Table tab with the raw source data for custom filtering.

    Values are unchanged; only COLUMN ORDER is rearranged for usability: Opp #,
    Revision #, and Property Name lead, followed by a report-like order, with any
    remaining ("extra") source columns appended at the end.
    """
    ws = wb.create_sheet("Table")
    ws.sheet_properties.tabColor = HERON
    ws.sheet_view.showGridLines = True

    preferred = [
        'Opportunity #', 'Revision #', 'Property Name',
        'Opportunity Name', 'Master Opportunity Name',
        'Company Name', 'Sales Rep', 'Operations Mgr Name',
        'Start Date', 'Oppty Complete Date', 'Won Date',
        'Branch', 'Division', 'Invoice Type', 'Opportunity Type',
        'Job Status', 'Opportunity Status Name',
        'Revenue Estimated', 'Earned Revenue', 'Invoiced Revenue',
        'Labor Hours Actual', 'Labor Hours Estimated', 'Future Scheduled Hours',
        'Labor Cost Actual', 'Labor Cost Estimated',
        'Material Cost Actual', 'Material Cost Estimated',
        'Sub Cost Actual', 'Sub Cost Estimated',
        'Equipment Cost Actual', 'Equipment Cost Estimated',
        'Other Cost Actual', 'Other Cost Estimated',
    ]
    ordered_cols = [c for c in preferred if c in raw_df.columns] + \
                   [c for c in raw_df.columns if c not in preferred]
    tdf = raw_df[ordered_cols]

    # Write headers
    for ci, col_name in enumerate(ordered_cols, 1):
        apply_cell(ws, 1, ci, col_name, font=font_body(10, bold=True, color=WHITE_HEX),
                   fill=FILL_HEADER, alignment=ALIGN_CENTER, border=THIN_BORDER)
        ws.column_dimensions[get_column_letter(ci)].width = 15

    # Write all data rows — values unchanged, reordered columns only
    for ri, (_, row_data) in enumerate(tdf.iterrows()):
        r = 2 + ri
        fill = FILL_ZEBRA if ri % 2 == 0 else FILL_WHITE
        for ci, col_name in enumerate(ordered_cols, 1):
            val = row_data[col_name]
            if isinstance(val, pd.Timestamp):
                val = val.to_pydatetime()
                nf = FMT_DATE
            elif isinstance(val, (int, float)) and not (isinstance(val, float) and np.isnan(val)):
                nf = None
            else:
                if isinstance(val, float) and np.isnan(val):
                    val = ''
                nf = None
            apply_cell(ws, r, ci, val, font=font_body(9), fill=fill,
                       alignment=ALIGN_LEFT, number_format=nf, border=THIN_BORDER)

    last_col = get_column_letter(len(ordered_cols))
    last_row = 1 + len(tdf)
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"
    ws.freeze_panes = 'D2'  # freeze through Property Name
    print(f"  Table tab: {len(tdf)} rows x {len(ordered_cols)} columns (reordered raw data)")


def build_version_history(wb, change_note="Report generated", user="Trey"):
    """Create or append to Version History tab (thg-report-standards Standard 9)."""
    tz = pytz.timezone("America/Chicago")
    now = datetime.now(tz)

    if "Version History" in wb.sheetnames:
        ws = wb["Version History"]
        next_row = ws.max_row + 1
    else:
        ws = wb.create_sheet("Version History")
        ws.sheet_properties.tabColor = TAB_COLORS["Version History"]
        ws.sheet_view.showGridLines = False

        ws.merge_cells("A1:C1")
        apply_cell(ws, 1, 1, "Version History", font=font_heading(14, True, WHITE_HEX),
                   fill=FILL_NAVY, alignment=ALIGN_CENTER)

        for ci, label in enumerate(["Date", "User", "Change"], 1):
            apply_cell(ws, 3, ci, label, font=font_heading(11, True, WHITE_HEX),
                       fill=FILL_NAVY, alignment=ALIGN_CENTER, border=THIN_BORDER)

        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 50
        next_row = 4

    ri = next_row - 4
    fill = FILL_ZEBRA if ri % 2 == 0 else FILL_WHITE
    apply_cell(ws, next_row, 1, _fmt_dt(now, with_time=True),
               font=font_body(10), fill=fill, alignment=ALIGN_CENTER, border=THIN_BORDER)
    apply_cell(ws, next_row, 2, user, font=font_body(10), fill=fill,
               alignment=ALIGN_LEFT_NOWRAP, border=THIN_BORDER)
    apply_cell(ws, next_row, 3, change_note, font=font_body(10), fill=fill,
               alignment=ALIGN_LEFT_NOWRAP, border=THIN_BORDER)

    # Ensure last tab
    target = len(wb.sheetnames) - 1
    current = wb.sheetnames.index("Version History")
    if current != target:
        wb.move_sheet("Version History", offset=target - current)


# ═══════════════════════════════════════════════════════════════════════════════
# RECONCILIATION (thg-report-standards Standard 4)
# ═══════════════════════════════════════════════════════════════════════════════

def run_reconciliation(ip_grouped, comp_grouped, dashboard_earned):
    print("\n═══ POST-BUILD RECONCILIATION ═══")
    ip_count = len(ip_grouped)
    comp_count = len(comp_grouped)
    print(f"  In Process grouped rows:   {ip_count}")
    print(f"  Completed grouped rows:    {comp_count}")
    print(f"  Total output rows:         {ip_count + comp_count}")

    all_pass = True
    if ip_count > 0:
        detail_earned = ip_grouped['Earned Revenue'].sum()
        match = abs(detail_earned - dashboard_earned) < 0.01
        status = "✅ PASS" if match else "❌ FAIL"
        print(f"  IP Earned Revenue tie-out:  {status}")
        print(f"    Detail: ${detail_earned:,.0f}  |  Dashboard: ${dashboard_earned:,.0f}")
        if not match: all_pass = False

    print(f"  Category exhaustion:       ✅ PASS")
    print(f"═══ RECONCILIATION {'PASSED' if all_pass else 'ISSUES FOUND'} ═══\n")
    return all_pass


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def parse_args():
    p = argparse.ArgumentParser(description="Build THG CIP Report")
    p.add_argument("input_file", help="Path to Aspire Opportunity export (.xlsx)")
    p.add_argument("output_file", help="Output Excel file path")
    p.add_argument("--client-name", required=True, help="Client display name")
    p.add_argument("--branch", action="append", dest="branches", help="Branch filter (repeatable; omit for all)")
    p.add_argument("--division", action="append", dest="divisions", required=True, help="Division filter (repeatable)")
    # Backend logic only — NOT prompted, NOT shown on the disclosure tab.
    p.add_argument("--original-opp-cutoff", default="2024-01-01",
                   help="Won Date cutoff (YYYY-MM-DD) for the Original Opp flag. Backend default "
                        "2024-01-01. Change-order revisions whose original opp was won before this "
                        "date get estimated costs zeroed. Override only if explicitly requested.")
    p.add_argument("--sub-margin", type=float, default=0.281,
                   help="Expected sub margin used to back into implied sub revenue on the Complete "
                        "Overview Non Sub Rev column (default 0.281 = 28.1%%). "
                        "Non Sub Rev = Earned Revenue - (Act Sub Cost / (1 - sub_margin)).")
    # ── Completed tab window (jobs shown on Completed tab + Completed Dashboard View) ──
    p.add_argument("--completed-range", choices=["this_month", "last_30_days", "ytd"],
                   default="this_month",
                   help="Completed-tab window, ending on the current day (not month-end). "
                        "this_month | last_30_days | ytd. Default this_month.")
    p.add_argument("--completed-start", default=None,
                   help="Custom completed-tab window start (YYYY-MM-DD). Overrides --completed-range.")
    p.add_argument("--completed-end", default=None,
                   help="Custom completed-tab window end (YYYY-MM-DD). Defaults to current day.")
    # ── Complete Overview window (bucket concentration analysis) — separate from above ──
    p.add_argument("--overview-range", choices=["last_year", "last_12_complete_months"],
                   default="last_12_complete_months",
                   help="Complete Overview window. last_year = prior calendar year; "
                        "last_12_complete_months = 12 full months ending last month. "
                        "Default last_12_complete_months.")
    # ── Invoice % highlight rule (In Process tabs) ──
    p.add_argument("--invoice-flag-rule", choices=["lag", "over"], default="lag",
                   help="In Process Invoice %% highlight. 'lag' = flag when Invoice %% is more than "
                        "--invoice-flag-gap below Rev %% Completed (default). 'over' = flag any "
                        "Invoice %% greater than Earned (Rev %% Completed).")
    p.add_argument("--invoice-flag-gap", type=float, default=0.10,
                   help="Gap below Rev %% Completed that triggers the Invoice %% highlight when "
                        "--invoice-flag-rule=lag (default 0.10 = 10%%).")
    # ── Dynamic Act/Est cost-pace threshold (In Process tabs, editable in-sheet on row 4) ──
    p.add_argument("--cost-pace-threshold", type=float, default=0.0,
                   help="Initial buffer above Rev %% Completed before an Act/Est column flags red on "
                        "the In Process tabs (default 0.0). Written to an editable cell on row 4; the "
                        "conditional formatting reads that cell so the user can change it live in Excel.")
    # ── Column selection (pipe-delimited header names; omit a flag to include all) ──
    p.add_argument("--ip-columns", default=None,
                   help="Pipe-delimited In Process headers to include. Omit to include all.")
    p.add_argument("--bp-columns", default=None,
                   help="Pipe-delimited In Process By Property headers to include. Omit for all.")
    p.add_argument("--comp-columns", default=None,
                   help="Pipe-delimited Completed headers to include. Omit for all.")
    p.add_argument("--user", default="Trey", help="User name for Version History")
    p.add_argument("--change-note", default="Report generated", help="Change note for Version History")
    p.add_argument("--logo", default=None, help="Path to THG logo PNG")
    p.add_argument("--no-logo", action="store_true", help="Skip logo insertion")
    p.add_argument("--min-est-revenue", type=float, default=0,
                   help="Minimum Revenue Estimated threshold (exclude opps below this amount)")
    return p.parse_args()


def main():
    args = parse_args()

    logo_path = None if args.no_logo else args.logo
    cutoff = pd.Timestamp(args.original_opp_cutoff)
    run_date = run_date_str()

    # ── Load ──
    print(f"Loading {args.input_file}...")
    raw = pd.read_excel(args.input_file)
    print(f"  Raw rows: {len(raw)}")

    # Parse dates
    for dc in ['Won Date', 'Oppty Complete Date', 'Start Date']:
        if dc in raw.columns:
            raw[dc] = pd.to_datetime(raw[dc], errors='coerce')

    # ── Validate required columns ──
    required = [
        'Branch', 'Opportunity #', 'Property Name', 'Division', 'Invoice Type',
        'Opportunity Name', 'Opportunity Type', 'Revision #', 'Job Status', 'Won Date',
        'Revenue Estimated', 'Earned Revenue', 'Invoiced Revenue',
        'Labor Hours Actual', 'Labor Hours Estimated',
        'Labor Cost Actual', 'Labor Cost Estimated', 'Material Cost Actual', 'Material Cost Estimated',
        'Sub Cost Actual', 'Sub Cost Estimated', 'Equipment Cost Actual', 'Equipment Cost Estimated',
        'Other Cost Actual', 'Other Cost Estimated',
    ]
    missing = [c for c in required if c not in raw.columns]
    if missing:
        print(f"ERROR: Missing required columns: {missing}", file=sys.stderr)
        sys.exit(1)

    # ── Pipeline ──
    print("Building Opp Master...")
    opp_master = build_opp_master(raw)
    print(f"  Opp Master rows: {len(opp_master)}")

    print("Building data queries...")
    (ip_grouped, ip_byprop, comp_grouped, overview_grouped,
     ip_grouped_all, dq, _c_start) = build_data_queries(
        raw, opp_master, args.divisions, args.branches, cutoff,
        completed_range=args.completed_range, overview_range=args.overview_range,
        min_est_revenue=args.min_est_revenue,
        completed_start=args.completed_start, completed_end=args.completed_end)
    print(f"  Filtered data rows: {len(dq)}")

    (ip_grouped, ip_byprop, comp_grouped, overview_grouped,
     ip_grouped_all) = finalize_datasets(
        ip_grouped, ip_byprop, comp_grouped, overview_grouped, ip_grouped_all)
    print(f"  In Process:    {len(ip_grouped)} jobs")
    print(f"  By Property:   {len(ip_byprop)} properties")
    print(f"  Completed:     {len(comp_grouped)} jobs (Completed tab)")
    print(f"  Overview:      {len(overview_grouped)} jobs (Complete Overview)")

    # Human-readable window labels for headers/disclosure
    completed_label = {
        'this_month': 'This month (to current day)',
        'last_30_days': 'Last 30 days',
        'ytd': 'Year to date',
    }.get(args.completed_range, args.completed_range)
    if args.completed_start:
        completed_label = f"Custom {args.completed_start} → {args.completed_end or 'today'}"
    overview_label = {
        'last_year': 'Last calendar year',
        'last_12_complete_months': 'Last 12 completed months',
    }.get(args.overview_range, args.overview_range)

    # ── Column selection (omit a flag to include all) ──
    def _select(cols, raw_spec):
        if not raw_spec:
            return cols
        wanted = [h.strip() for h in raw_spec.split('|') if h.strip()]
        filtered = [c for c in cols if c[0] in wanted]
        return filtered if filtered else cols
    ip_cols = _select(IP_COLS, args.ip_columns)
    bp_cols = _select(BP_COLS, args.bp_columns)
    comp_cols = _select(COMP_COLS, args.comp_columns)

    # ── Build workbook ──
    print("Building workbook...")
    wb = Workbook()

    build_proprietary_disclosure(wb, args.client_name, logo_path, run_date,
                                 args.branches, args.divisions,
                                 completed_label, overview_label,
                                 sub_margin=args.sub_margin)

    dashboard_earned = build_dashboard(wb, args.client_name, logo_path, run_date,
                                       ip_grouped, comp_grouped, args.divisions,
                                       completed_label, ip_grouped_all=ip_grouped_all)

    build_data_tab(wb, "In Process", ip_cols, ip_grouped, args.client_name, logo_path, run_date,
                   invoice_flag_rule=args.invoice_flag_rule, invoice_flag_gap=args.invoice_flag_gap,
                   cost_pace_threshold=args.cost_pace_threshold)
    build_data_tab(wb, "In Process By Property", bp_cols, ip_byprop, args.client_name, logo_path, run_date,
                   invoice_flag_rule=args.invoice_flag_rule, invoice_flag_gap=args.invoice_flag_gap,
                   cost_pace_threshold=args.cost_pace_threshold)
    build_data_tab(wb, "Completed", comp_cols, comp_grouped, args.client_name, logo_path, run_date)

    build_completed_dashboard_view(wb, args.client_name, logo_path, run_date, comp_grouped)

    build_complete_overview(wb, args.client_name, logo_path, run_date,
                            overview_grouped, DEFAULT_GM_BUCKETS, DEFAULT_REV_BUCKETS,
                            sub_margin=args.sub_margin)

    build_table_tab(wb, raw)

    build_version_history(wb, change_note=args.change_note, user=args.user)

    # ── Save ──
    print(f"Saving to {args.output_file}...")
    wb.save(args.output_file)

    # ── Reconciliation (Standard 4) ──
    recon_ok = run_reconciliation(ip_grouped, comp_grouped, dashboard_earned)

    if not recon_ok:
        print("WARNING: Reconciliation issues detected. Review before delivering.", file=sys.stderr)
        sys.exit(1)

    print("Done.")
    sys.exit(0)


if __name__ == "__main__":
    main()
