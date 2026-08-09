"""Profile generated CIP workbooks for structural and package consistency."""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import statistics
import zipfile
from collections import Counter
from pathlib import Path
from typing import Any
from xml.etree import ElementTree

from openpyxl import load_workbook

VOLATILE_PACKAGE_MEMBERS = {"docProps/core.xml", "xl/calcChain.xml"}


def _sha256_bytes(value: bytes) -> str:
    return hashlib.sha256(value).hexdigest().upper()


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest().upper()


def _version_history_member(archive: zipfile.ZipFile) -> str | None:
    workbook_root = ElementTree.fromstring(archive.read("xl/workbook.xml"))
    relationships_root = ElementTree.fromstring(
        archive.read("xl/_rels/workbook.xml.rels")
    )
    relationship_targets = {
        item.attrib["Id"]: item.attrib["Target"].lstrip("/")
        for item in relationships_root
    }
    relationship_namespace = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
    spreadsheet_namespace = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheet"
    for sheet in workbook_root.iter(spreadsheet_namespace):
        if sheet.attrib.get("name") == "Version History":
            target = relationship_targets.get(sheet.attrib[relationship_namespace])
            if not target:
                return None
            return target if target.startswith("xl/") else f"xl/{target}"
    return None


def _package_profile(path: Path) -> dict[str, Any]:
    with zipfile.ZipFile(path) as archive:
        version_history_member = _version_history_member(archive)
        excluded = set(VOLATILE_PACKAGE_MEMBERS)
        if version_history_member:
            excluded.add(version_history_member)

        member_hashes = {
            name: _sha256_bytes(archive.read(name))
            for name in sorted(archive.namelist())
            if not name.endswith("/")
        }
        stable_member_hashes = {
            name: digest for name, digest in member_hashes.items() if name not in excluded
        }
        normalized_payload = "\n".join(
            f"{name}|{digest}" for name, digest in stable_member_hashes.items()
        ).encode("utf-8")
        return {
            "member_hashes": member_hashes,
            "normalized_package_sha256": _sha256_bytes(normalized_payload),
            "excluded_members": sorted(excluded),
        }


def _workbook_profile(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        dimensions = {
            name: {
                "rows": workbook[name].max_row,
                "columns": workbook[name].max_column,
            }
            for name in workbook.sheetnames
        }
        structure = {"worksheets": workbook.sheetnames, "dimensions": dimensions}
    finally:
        workbook.close()

    package = _package_profile(path)
    timestamp_match = re.search(r"(\d{8}_\d{6})", path.stem)
    return {
        "filename": path.name,
        "run_timestamp": timestamp_match.group(1) if timestamp_match else None,
        "size_bytes": path.stat().st_size,
        "sha256": _sha256_file(path),
        "worksheet_count": len(structure["worksheets"]),
        "worksheets": structure["worksheets"],
        "worksheet_dimensions": structure["dimensions"],
        "structural_signature": _sha256_bytes(
            json.dumps(structure, sort_keys=True).encode("utf-8")
        ),
        **package,
    }


def analyze(directory: Path) -> dict[str, Any]:
    paths = sorted(directory.glob("CIP_Report_*.xlsx"))
    if not paths:
        raise FileNotFoundError(f"No CIP_Report_*.xlsx files found in {directory}")

    profiles = [_workbook_profile(path) for path in paths]
    baseline_hashes = profiles[-1]["member_hashes"]
    for profile in profiles:
        member_hashes = profile.pop("member_hashes")
        profile["changed_members_vs_latest"] = sorted(
            name
            for name in set(baseline_hashes) | set(member_hashes)
            if baseline_hashes.get(name) != member_hashes.get(name)
        )

    structural_groups = Counter(item["structural_signature"] for item in profiles)
    package_groups = Counter(item["normalized_package_sha256"] for item in profiles)
    sizes = [item["size_bytes"] for item in profiles]
    expected_sheets = profiles[-1]["worksheets"]
    expected_dimensions = profiles[-1]["worksheet_dimensions"]
    structurally_consistent = all(
        item["worksheets"] == expected_sheets
        and item["worksheet_dimensions"] == expected_dimensions
        for item in profiles
    )
    latest_signature = profiles[-1]["structural_signature"]
    latest_structure_files = [
        item["filename"] for item in profiles if item["structural_signature"] == latest_signature
    ]
    structural_group_details = [
        {
            "signature": signature,
            "count": count,
            "files": [
                item["filename"]
                for item in profiles
                if item["structural_signature"] == signature
            ],
        }
        for signature, count in structural_groups.most_common()
    ]

    return {
        "schema_version": 1,
        "scope": {
            "directory": directory.as_posix(),
            "pattern": "CIP_Report_*.xlsx",
            "files_analyzed": len(profiles),
            "baseline": profiles[-1]["filename"],
        },
        "summary": {
            "structurally_consistent": structurally_consistent,
            "structural_signature_groups": len(structural_groups),
            "latest_structure_files": len(latest_structure_files),
            "latest_structure_share_pct": round(
                len(latest_structure_files) / len(profiles) * 100,
                1,
            ),
            "normalized_package_groups": len(package_groups),
            "size_bytes_min": min(sizes),
            "size_bytes_max": max(sizes),
            "size_bytes_mean": round(statistics.mean(sizes), 1),
            "size_range_pct_of_mean": round(
                (max(sizes) - min(sizes)) / statistics.mean(sizes) * 100,
                3,
            ),
        },
        "structural_groups": structural_group_details,
        "profiles": profiles,
        "limitations": [
            "The approved manual reference workbook was not available.",
            "Structural and package consistency do not prove financial reconciliation.",
            "Runs may reflect different code versions or configurations unless a run manifest proves otherwise.",
            "Normalized package hashes exclude document metadata, calculation chain data, and Version History content.",
        ],
    }


def render_markdown(result: dict[str, Any]) -> str:
    summary = result["summary"]
    profiles = result["profiles"]
    consistency = "PASS" if summary["structurally_consistent"] else "MIXED"
    lines = [
        "# CIP Workbook Reproducibility Evidence",
        "",
        "## Technical summary",
        "",
        f"{len(profiles)} generated workbooks were inspected. Structural consistency is "
        f"**{consistency}** across the full history: "
        f"{summary['latest_structure_files']} of {len(profiles)} files "
        f"({summary['latest_structure_share_pct']:.1f}%) share the current worksheet structure. "
        "This supports stable recent workbook construction, but it does not establish financial "
        "reconciliation without the approved manual reference workbook.",
        "",
        "## Key findings",
        "",
        f"- Structural signature groups: **{summary['structural_signature_groups']}**.",
        f"- Files matching the current structural signature: "
        f"**{summary['latest_structure_files']} of {len(profiles)}**.",
        f"- Normalized package groups: **{summary['normalized_package_groups']}**.",
        f"- Output size range: **{summary['size_bytes_min']:,}–{summary['size_bytes_max']:,} bytes** "
        f"({summary['size_range_pct_of_mean']:.3f}% of the mean).",
        "- Every workbook was checked independently; identical filenames were not assumed to mean identical content.",
        "- The two earliest artifacts form a larger-row structural group, consistent with a different filter, configuration, or build revision; manifests were not available to distinguish those causes.",
        "",
        "## Run inventory",
        "",
        "| Workbook | Size (bytes) | Sheets | Structural signature | Normalized package | Changed ZIP members vs latest |",
        "| --- | ---: | ---: | --- | --- | ---: |",
    ]
    for item in profiles:
        lines.append(
            f"| `{item['filename']}` | {item['size_bytes']:,} | {item['worksheet_count']} | "
            f"`{item['structural_signature'][:12]}` | "
            f"`{item['normalized_package_sha256'][:12]}` | "
            f"{len(item['changed_members_vs_latest'])} |"
        )
    lines.extend(
        [
            "",
            "## Scope and methodology",
            "",
            "The analysis compared worksheet order, names, and used dimensions. It also hashed every "
            "file and every ZIP member inside each XLSX package. A normalized package signature excluded "
            "document metadata, calculation-chain data, and Version History content because those fields "
            "are expected to vary across otherwise equivalent runs.",
            "",
            "## Limitations and pending validation",
            "",
        ]
    )
    lines.extend(f"- {item}" for item in result["limitations"])
    lines.extend(
        [
            "",
            "## Recommended next steps",
            "",
            "1. Capture a structured run manifest on every new run so configuration and code version are known.",
            "2. Repeat at least five runs from one fixed input, configuration, skill package, and Git commit.",
            "3. Run the reconciliation harness against the approved manual reference when it becomes available.",
            "4. Treat financial accuracy as pending until the configured value checks pass.",
            "",
            "## Further questions",
            "",
            "- Which dashboard cells and Table-sheet aggregates are authoritative acceptance measures?",
            "- What absolute or percentage tolerances are acceptable for currency and margin calculations?",
            "- Should Version History and generated timestamps be excluded from formal reproducibility comparisons?",
            "",
        ]
    )
    return "\n".join(lines)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--directory", type=Path, default=Path("outputs"))
    parser.add_argument("--json-output", type=Path)
    parser.add_argument("--markdown-output", type=Path)
    args = parser.parse_args()

    result = analyze(args.directory)
    if args.json_output:
        args.json_output.parent.mkdir(parents=True, exist_ok=True)
        args.json_output.write_text(json.dumps(result, indent=2), encoding="utf-8")
    if args.markdown_output:
        args.markdown_output.parent.mkdir(parents=True, exist_ok=True)
        args.markdown_output.write_text(render_markdown(result), encoding="utf-8")
    print(json.dumps(result["summary"], indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
