"""Compare a generated workbook with an approved reference workbook."""

from __future__ import annotations

import argparse
import json
import math
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


def compare_values(
    reference: Any,
    candidate: Any,
    *,
    absolute_tolerance: float = 0.0,
    relative_tolerance: float = 0.0,
) -> dict[str, Any]:
    if isinstance(reference, (int, float)) and isinstance(candidate, (int, float)):
        reference_number = float(reference)
        candidate_number = float(candidate)
        difference = candidate_number - reference_number
        allowed = max(
            absolute_tolerance,
            abs(reference_number) * relative_tolerance,
        )
        passed = math.isclose(
            candidate_number,
            reference_number,
            rel_tol=relative_tolerance,
            abs_tol=absolute_tolerance,
        )
        return {
            "reference": reference,
            "candidate": candidate,
            "difference": difference,
            "allowed_difference": allowed,
            "passed": passed,
        }
    return {
        "reference": reference,
        "candidate": candidate,
        "difference": None,
        "allowed_difference": None,
        "passed": reference == candidate,
    }


def _aggregate(values: Iterable[Any], operation: str) -> float | int:
    materialized = list(values)
    if operation == "count":
        return len(materialized)
    if operation == "nonblank_count":
        return sum(value not in (None, "") for value in materialized)
    if operation == "sum":
        return sum(float(value) for value in materialized if isinstance(value, (int, float)))
    raise ValueError(f"Unsupported aggregate operation: {operation}")


def _column_values(sheet, *, header_row: int, column_name: str) -> list[Any]:
    headers = [cell.value for cell in sheet[header_row]]
    try:
        column_index = headers.index(column_name) + 1
    except ValueError as exc:
        raise ValueError(
            f"Column {column_name!r} not found on {sheet.title!r} row {header_row}"
        ) from exc
    return [
        sheet.cell(row=row_index, column=column_index).value
        for row_index in range(header_row + 1, sheet.max_row + 1)
    ]


def _check_value(workbook, check: dict[str, Any]) -> Any:
    sheet = workbook[check["sheet"]]
    check_type = check["type"]
    if check_type == "cell":
        return sheet[check["cell"]].value
    if check_type == "column_aggregate":
        values = _column_values(
            sheet,
            header_row=int(check.get("header_row", 1)),
            column_name=check["column"],
        )
        return _aggregate(values, check["aggregate"])
    raise ValueError(f"Unsupported check type: {check_type}")


def reconcile(
    reference_path: Path,
    candidate_path: Path,
    spec: dict[str, Any] | None = None,
) -> dict[str, Any]:
    settings = spec or {}
    reference = load_workbook(reference_path, read_only=True, data_only=False)
    candidate = load_workbook(candidate_path, read_only=True, data_only=False)
    try:
        reference_sheets = reference.sheetnames
        candidate_sheets = candidate.sheetnames
        required_sheets = settings.get("required_sheets") or reference_sheets
        missing_reference = [name for name in required_sheets if name not in reference_sheets]
        missing_candidate = [name for name in required_sheets if name not in candidate_sheets]

        structural_checks: list[dict[str, Any]] = [
            {
                "id": "required_worksheets",
                "passed": not missing_reference and not missing_candidate,
                "missing_from_reference": missing_reference,
                "missing_from_candidate": missing_candidate,
            }
        ]
        if settings.get("compare_dimensions", True):
            for sheet_name in required_sheets:
                if sheet_name not in reference_sheets or sheet_name not in candidate_sheets:
                    continue
                expected = {
                    "rows": reference[sheet_name].max_row,
                    "columns": reference[sheet_name].max_column,
                }
                observed = {
                    "rows": candidate[sheet_name].max_row,
                    "columns": candidate[sheet_name].max_column,
                }
                structural_checks.append(
                    {
                        "id": f"dimensions:{sheet_name}",
                        "sheet": sheet_name,
                        "reference": expected,
                        "candidate": observed,
                        "passed": expected == observed,
                    }
                )

        value_checks: list[dict[str, Any]] = []
        defaults = settings.get("tolerances") or {}
        for check in settings.get("checks") or []:
            result = {
                "id": check["id"],
                "type": check["type"],
                "sheet": check["sheet"],
            }
            try:
                reference_value = _check_value(reference, check)
                candidate_value = _check_value(candidate, check)
                result.update(
                    compare_values(
                        reference_value,
                        candidate_value,
                        absolute_tolerance=float(
                            check.get(
                                "absolute_tolerance",
                                defaults.get("absolute", 0.0),
                            )
                        ),
                        relative_tolerance=float(
                            check.get(
                                "relative_tolerance",
                                defaults.get("relative", 0.0),
                            )
                        ),
                    )
                )
            except (KeyError, TypeError, ValueError) as exc:
                result.update({"passed": False, "error": str(exc)})
            value_checks.append(result)

        all_checks = structural_checks + value_checks
        return {
            "schema_version": 1,
            "overall_result": "PASS" if all(item["passed"] for item in all_checks) else "FAIL",
            "reference": reference_path.name,
            "candidate": candidate_path.name,
            "structural_checks": structural_checks,
            "value_checks": value_checks,
            "summary": {
                "checks": len(all_checks),
                "passed": sum(bool(item["passed"]) for item in all_checks),
                "failed": sum(not bool(item["passed"]) for item in all_checks),
            },
            "limitations": (
                []
                if value_checks
                else [
                    "No financial value checks were configured; this result validates structure only."
                ]
            ),
        }
    finally:
        reference.close()
        candidate.close()


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("reference", type=Path)
    parser.add_argument("candidate", type=Path)
    parser.add_argument("--spec", type=Path, help="JSON reconciliation specification")
    parser.add_argument("--output", type=Path, help="Optional JSON result path")
    args = parser.parse_args()

    spec = json.loads(args.spec.read_text(encoding="utf-8")) if args.spec else None
    result = reconcile(args.reference, args.candidate, spec)
    rendered = json.dumps(result, indent=2)
    if args.output:
        args.output.parent.mkdir(parents=True, exist_ok=True)
        args.output.write_text(rendered, encoding="utf-8")
    print(rendered)
    return 0 if result["overall_result"] == "PASS" else 1


if __name__ == "__main__":
    raise SystemExit(main())
