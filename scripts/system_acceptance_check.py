"""Acceptance checks for a generated CIP workbook and the live runner API."""

from __future__ import annotations

import argparse
import json
import urllib.error
import urllib.request
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


def get_json(url: str) -> dict:
    with urllib.request.urlopen(url, timeout=10) as response:
        return json.load(response)


def post_json(url: str, payload: dict) -> dict:
    request = urllib.request.Request(
        url,
        data=json.dumps(payload).encode("utf-8"),
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    with urllib.request.urlopen(request, timeout=30) as response:
        return json.load(response)


def check_workbook(path: Path) -> dict:
    if not path.is_file():
        raise AssertionError(f"Workbook does not exist: {path}")
    if path.stat().st_size < 100_000:
        raise AssertionError("Workbook is unexpectedly small")

    workbook = load_workbook(path, read_only=True, data_only=False)
    sheet_names = workbook.sheetnames
    required_sheets = {
        "Dashboard",
        "In Process",
        "Completed",
        "Complete Overview",
        "Table",
        "Version History",
    }
    missing_sheets = sorted(required_sheets.difference(sheet_names))
    if missing_sheets:
        raise AssertionError(f"Missing required sheets: {missing_sheets}")

    nonempty_sheets = {}
    for name in sheet_names:
        sheet = workbook[name]
        nonempty_sheets[name] = {"rows": sheet.max_row, "columns": sheet.max_column}
        if sheet.max_row < 1 or sheet.max_column < 1:
            raise AssertionError(f"Sheet is empty: {name}")
    workbook.close()

    return {
        "path": str(path.resolve()),
        "size_bytes": path.stat().st_size,
        "sheet_names": sheet_names,
        "sheet_dimensions": nonempty_sheets,
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--base-url", default="http://127.0.0.1:8787")
    args = parser.parse_args()

    checks = []
    try:
        health = get_json(f"{args.base_url}/health")
        assert health == {"status": "ok"}, health
        checks.append({"id": "AT-01", "name": "API health", "result": "PASS"})

        data = get_json(f"{args.base_url}/data")
        assert isinstance(data.get("files"), list), data
        checks.append({"id": "AT-02", "name": "Data inventory", "result": "PASS"})

        workbook = check_workbook(args.workbook)
        checks.append({"id": "AT-03", "name": "Workbook opens and required sheets exist", "result": "PASS"})

        try:
            post_json(f"{args.base_url}/inspect", {"filename": "../outside.xlsx"})
            raise AssertionError("Traversal request unexpectedly succeeded")
        except urllib.error.HTTPError as exc:
            assert exc.code == 404, exc.code
        checks.append({"id": "AT-04", "name": "Invalid request rejected", "result": "PASS"})

        report = {
            "timestamp_utc": datetime.now(timezone.utc).isoformat(),
            "overall_result": "PASS",
            "checks": checks,
            "workbook": workbook,
        }
        print(json.dumps(report, indent=2))
        return 0
    except Exception as exc:
        checks.append({"id": "AT-FAIL", "name": type(exc).__name__, "result": "FAIL", "detail": str(exc)})
        print(json.dumps({"overall_result": "FAIL", "checks": checks}, indent=2))
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
